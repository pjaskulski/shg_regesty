""" tworzenie list regestów ze Słownika Historyczno-Geograficznego """

import re
from pathlib import Path
import openpyxl
from tools import create_sheet, get_regesty, get_year_from_regest
from tools import get_source_from_regest, get_contents_from_regest


slowniki = {1: 'Benedyktyni',
            2: 'Chełmno',
            3: 'Kraków',
            4: 'Lublin',
            5: 'Lublin_zaginione',
            6: 'Płock',
            7: 'Poznań',
            8: 'Sanok',
            9: 'Wieluń',
            10: 'Wyszogród',
            11: 'Warszawa',
            12: 'Liw'
}

lista = {
            r'własn\.\s+szlach\.':'szlachecka',
            r'własn\.\s+kl\.':'klasztorna',
            r'własn\.\s+opata':'opata',
            r'własn\.\s+konw.':'konwentu',
            r'własn\.\s+ryc\.':'rycerska',
            r'należy\s+do\s+stołu\s+konw\.':'konwentu',
            r'należy\s+do\s+bpa\s+lubus\.':'biskupia',
            r'własn\.\s+podzielona\.':'podzielona',
            r'wieś\s+szlach\.':'szlachecka',
            r'należy\s+do\s+kl\.':'klasztorna',
            r'własn\.\s+wwdy':'wojewody',
            r'należy\s+do\s+[a-zA-ząśężźćńłó]+\s+stołu\s+opata':'opata',
            r'należy\s+do\s+stołu\s+opata':'opata',
            r'własn\.\s+rządu':'rządowa',
            r'własn\.\s+król\.':'królewska',
            r'należy\s+do\s+króla':'królewska',
            r'należy\s+do\s+opactwa':'opata',
            r'własn\.\s+szlach\.\s+i\s+prep\.':'szlachecka, prepozyta',
            r'własn\.\s+bpa':'biskupia',
            r'własn\.\s+szlach\.,\s+następnie\s+kl\.':'szlachecka, klasztorna',
            r'własn\.\s+bpa[a-zA-Z\.\s]+,\s+następnie\s+rządowa':'biskupia, rządowa',
            r'własn. kl.[a-zA-Ząśężźźćńłó\.,\s]+pleb\.,\s+[a-zA-Zśężźćńłó]\s+szlach\.':'klasztorna, plebana, szlachecka',
            r'własn\.\s+kl\.[a-zA-Ząśężźźćńłó\.,\s]+prep\.':'klasztorna, prepozyta',
            r'własn\.\s+bp':'biskupia',
            r'własn\.\s+arcbpstwa':'arcybiskupia',
            r'własn\.\s+książęca':'książęca',
            r'własn\.\s+książęca[a-zA-Z1-9ąśężźźćńłó\.,\s]+\s+król\.':'książęca, królewska',
            r'/własn.\s+książęca[a-zA-Ząśężźćńłó,\s]+szlach\./gm':'książęca, szlachecka',
            r'własn\.\s+książąt':'książęca',
            r'własn\.\s+książęca[a-zA-Ząśężźćńłó,\s]+kl\.':'książęca, klasztorna'
        }


def remove_html(value: str) -> str:
    # czyszczenie z pozostałości HTML-a
    pattern = r'<.*?>'
    value = re.sub(pattern, '', value)
    value = value.replace(r'\n', ' ')
    return value


def process_point(sheet, m_id, m_nazwa, point_num, point_text):
    """przetwarzanie treści punktu z zapisem do wskazanego arkusza
       sheet - wskazanie na arkusz
       m_id - identyfikator miejscowości
       nazwa - nazwa miejscowości
       point_num - numer punktu
       point_text - treść punktu
    """
    global regesty_count
    global place_owner

    reg_list = []
    point = point_num

    # if 'Własn. szlach. w kluczu melsztyńskim' in point_text:
    #     print()

    reg_list, ownership = get_regesty(point_text)

    reg_list = [remove_html(reg) for reg in reg_list]

    if reg_list:
        # pierwszy regest może wymagać podzielenia
        pattern = r'\.\s{1}\d{4}'
        pattern2 = r'\s+r\.\s{1}\d{4}'
        match = re.search(pattern, reg_list[0])
        match2 = re.search(pattern2, reg_list[0])
        if match and not match2:
            pos = reg_list[0].find('. 1')
            if pos != -1:
                part1 = reg_list[0][:pos + 1]
                part2 = reg_list[0][pos + 2:]
                reg_list[0] = part2
                reg_list.insert(0, part1)

    # własność
    if point_num == 3:
        for reg in reg_list:
            year = get_year_from_regest(reg)
            for pattern, owner_type in lista.items():
                match = re.search(pattern, reg.lower())
                if match:
                    if (m_id, year) in place_owner:
                        _m_id, _year, owner_types = place_owner[(m_id, year)]
                        if owner_type not in owner_types:
                            if not owner_types.endswith(','):
                                owner_types += ', '
                            owner_types += owner_type
                            place_owner[(m_id, year)] = (m_nazwa, year, owner_types)
                    else:
                        place_owner[(m_id, year)] = (m_nazwa, year, owner_type)

    # specjalna obsługa dla p. 1
    if point_num == 1:
        new_reg_list = []
        for item in reg_list:
            if ")," in item:
                tmp = item.split("),")
                tmp = [r+')' if not r.endswith(")") else r for r in tmp]
                new_reg_list = new_reg_list + tmp
            elif ")." in item:
                tmp = item.split(").")
                tmp = [r+')' if not r.endswith(")") else r for r in tmp]
                new_reg_list = new_reg_list + tmp
            else:
                new_reg_list.append(item)
        reg_list = new_reg_list

    for item in reg_list:
        # if 'n. par. Nowa Słupia (DLb. II 490) [' in item:
        #     print()
        item = item.strip()
        year = get_year_from_regest(item)
        source = get_source_from_regest(item)
        contents = get_contents_from_regest(item, year, source, m_nazwa)
        contents = contents.replace('[@', '[').replace('@]',']')
        item = item.replace('[@', '[').replace('@]',']')
        tmp_regest = [m_id, m_nazwa, point, year, contents, source, item]
        regest = tuple(tmp_regest)
        sheet.append(regest)
        regesty_count[point_num] += 1


# -------------- Główny program ------------------------------------------------
if __name__ == '__main__':
    regesty_count = {1:0, 2:0, 3:0, 4:0, 5:0, 6:0}

    for nr in range(1, 13):
        print(f"Słownik: {slowniki[nr]} ({nr})")

        place_owner = {}

        output_path = Path('.').parent / f'output/wyniki/regesty_{slowniki[nr]}.xlsx'
        input_path = Path('.').parent / f'output/wyniki/hasla_{nr}_pkt.xlsx'

        wb_hasla = openpyxl.load_workbook(input_path)
        ws_hasla = wb_hasla.active

        wb = openpyxl.Workbook()
        ws = wb.active

        point_sheet_columns = ['Miejscowość_Id', 'Nazwa', 'Punkt', 'Rok',
                               'Treść', 'Źródła', 'Cały Regest']
        regest_sheet_1 = create_sheet(wb=wb, title="Regesty p. 1", columns=point_sheet_columns)
        regest_sheet_2 = create_sheet(wb=wb, title="Regesty p. 2", columns=point_sheet_columns)
        regest_sheet_3 = create_sheet(wb=wb, title="Regesty p. 3", columns=point_sheet_columns)
        regest_sheet_4 = create_sheet(wb=wb, title="Regesty p. 4", columns=point_sheet_columns)
        regest_sheet_5 = create_sheet(wb=wb, title="Regesty p. 5", columns=point_sheet_columns)
        regest_sheet_6 = create_sheet(wb=wb, title="Regesty p. 6", columns=point_sheet_columns)

        # kolumny w źródłowym pierwszym arkuszu
        col_names = {}
        nr_col = 0
        for column in ws_hasla.iter_cols(1, ws_hasla.max_column):
            col_names[column[0].value] = nr_col
            nr_col += 1

        # przetwarzanie wierszy
        max_hasla = ws_hasla.max_row
        licznik = 0
        for row in ws_hasla.iter_rows(2, max_hasla):
            licznik += 1
            print(f'Przetwarzam wiersz {licznik} z {max_hasla}.')
            miejsce_id = row[col_names['id']].value
            nazwa = row[col_names['header']].value
            nazwa = re.sub(r'<.*?>', '', nazwa)
            p1 = row[col_names[1]].value
            p2 = row[col_names[2]].value
            p3 = row[col_names[3]].value
            p4 = row[col_names[4]].value
            p5 = row[col_names[5]].value
            p6 = row[col_names[6]].value

            # Punkt 1
            if p1 is not None:
                process_point(regest_sheet_1, miejsce_id, nazwa, 1, p1)
            if p2 is not None:
                process_point(regest_sheet_2, miejsce_id, nazwa, 2, p2)
            if p3 is not None:
                process_point(regest_sheet_3, miejsce_id, nazwa, 3, p3)
            if p4 is not None:
                process_point(regest_sheet_4, miejsce_id, nazwa, 4, p4)
            if p5 is not None:
                process_point(regest_sheet_5, miejsce_id, nazwa, 5, p5)
            if p6 is not None:
                process_point(regest_sheet_6, miejsce_id, nazwa, 6, p6)

        # usunięcie pustego pierwszego arkusza
        sheet1 = wb['Sheet']
        wb.remove(sheet1)

        # arkusz z własnością
        wlasnosc_sheet = create_sheet(wb=wb, title="Własność", columns=['Miejsc_ID', 'Miejscowość', 'Rok', 'Własność'])
        for key_tuple, owner_lista in place_owner.items():
            miejscowosc_id, _year = key_tuple
            miejscowosc, rok, owner_types = owner_lista
            
            # bez powtórzeń typów własności
            tmp = owner_types.split(',')
            if len(tmp) > 1:
                tmp = [t.strip() for t in tmp]
                tmp = list(set(tmp))
                owner_types = ', '.join(tmp)

            wiersz = (miejscowosc_id, miejscowosc, rok, owner_types)
            wlasnosc_sheet.append(wiersz)

        # zapis pliku xlsx
        wb.save(filename=output_path)

    suma = 0
    for i in range(1, 7):
        print(f"p.{i}", regesty_count[i])
        suma += regesty_count[i]
    print(f"razem: {suma}")
        