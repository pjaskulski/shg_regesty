""" tworzenie list regestów ze Słownika Historyczno-Geograficznego """

import re
import csv
import sys
import os
from pathlib import Path
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from tools import create_sheet


# mamy bardzo duże pola i bez zwiększenia limitu skrypt się wykłada
csv.field_size_limit(sys.maxsize)

slowniki = {1: 'Benedyktyni',
            2: 'Chełmno',
            3: 'Kraków',
            4: 'Lublin',
            5: 'Lublin_zaginione',
            6: 'Płock',
            7: 'Poznań',
            8: 'Sanok',
            9: 'Wieluń',
            10: 'Wyszogród',
            11: 'Warszawa',
            12: 'Liw',
            13: 'Czersk',
            14: 'Podlasie'
}


def all_caps(value: str) -> bool:
    """ czy nazwa pisana głównie dużymi literami? """
    result = True

    if value == 'IZDEBKI-Guzy':
        value = 'IZDEBKI-GUZY'
    if value == 'SIEDLCE m.':
        value = 'SIEDLCE M.'

    if '[' in value:
        pos = value.find('[')
        if pos != -1:
            value = value[:pos].strip()

    eliminacja = [' lub ', ' i ', ' alias ', ' albo ',
                  'przekazy nie zlokalizowane',
                  ' zw. też ', ' czyli ', ' następnie ',
                  'przekazy nie określone', ' zamek ',
                  'ziemia i gród', 'później', ' i wieś',
                  ' koło ']

    for element in eliminacja:
        if element in value:
            value = value.replace(element, ' ').strip()

    off = False
    lower_counter = 0
    for ch in value:
        if ch == '[':
            off = True
        elif ch == ']':
            off = False
        elif ch == '–':
            break
        elif not off and ch not in [' ', ',', ';', '.', '-', '–'] and ch.islower():
            lower_counter += 1

    #if lower_counter > len(nazwa) - lower_counter:
    if lower_counter > 0:
        result = False

    return result


def remove_html(value: str) -> str:
    """ czyszczenie z pozostałości HTML-a """
    # usuwanie przypisu z zawartością
    html_start = '<sup>'
    html_end = '</sup>'
    pos1 = value.find(html_start)
    if pos1 != -1:
        pos2 = value.find(html_end)
        if pos2 != -1:
            value = value[:pos1] + ' ' + value[pos2 + len(html_end):]

    html_start = '<span id="footnote'
    html_end = '</span>'
    pos1 = value.find(html_start)
    if pos1 != -1:
        pos2 = value.find(html_end)
        if pos2 != -1:
            value = value[:pos1] + ' ' + value[pos2 + len(html_end):]

    pattern = r'<.*?>'
    value = re.sub(pattern, '', value)
    value = value.replace(r'\n', ' ')
    return value

def double_space(value:str) -> str:
    """ usuwa podwójne spacje z przekazanego tekstu """
    return ' '.join(value.strip().split())

def shortcut(slownik: str) -> dict:
    """ skróty geograficzne
        slownik - nazwa słownika
        zwraca dict ze skrotami dla danego słownika
    """
    if slownik == 'Benedyktyni':
        shortcuts = {"gnieźn.":"gnieźnieński",
                    "kiel.":"kielecki",
                    "koz.":"kozienicki",
                    "król.":"królewski",
                    "lub.":"lubelski",
                    "lubus.":"lubuski",
                    "łysog.":"łysogórski",
                    "opat.":"opatowski",
                    "pap.":"papieski",
                    "rad.":"radomski",
                    "sand.":"sandomierski",
                    "sieciech.":"sieciechowski",
                    "sier.":"sieradzki",
                    "święt.":"świętokrzyski",
                    "tyn.":"tyniecki",
                    "urzęd.":"urzędowski",
                    "urząd.":"urzędowski",
                    "wąch.":"wąchocki",
                    "wąwol.":"wąwolnicki",
                    "wiśl.":"wiślicki",
                    "włocł.":"włocławski",
                    "krak.":"krakowski"
                    }
    elif slownik == 'Chełmno':
        shortcuts = {"bierzg.":"bierzgłowski",
                     "brat.":"bratiański",
                     "brod.":"brodnicki",
                     "chełm.":"chełmiński",
                     "dzierzg.":"dzierzgoński",
                     "gol.":"golubski",
                     "grudz.":"grudziądzki",
                     "kol.":"koloński",
                     "kow.":"kowalewski",
                     "krzyż.":"krzyżacki",
                     "kuj.":"kujawski",
                     "kurz.":"kurzętnicki",
                     "lip.":"lipieniecki",
                     "lub.":"lubawski",
                     "magd.":"magdeburski",
                     "maz.":"mazowiecki",
                     "mich.":"michałowski",
                     "miej.":"miejski",
                     "pap.":"papowski",
                     "płoc.":"płocki",
                     "pokrz.":"pokrzywnieński",
                     "pomez.":"pomezański",
                     "radz.":"radzyński",
                     "rog.":"rogoziński",
                     "starogr.":"starogrodzki",
                     "tor.":"toruński",
                     "unisł.":"unisławski",
                     "warm.":"warmiński",
                     "włocł.":"włocławski"
                     }
    elif slownik == 'Kraków':
        shortcuts = {"krak.":"krakowski",
                     "lel.":"lelowski",
                     "biec.":"biecki",
                     "chęc.":"chęciński",
                     "czchow.":"czchowski",
                     "czech.":"czechowski",
                     "czes.":"czeski",
                     "frank.":"frankoński",
                     "gnieźn.":"gnieźnieński",
                     "imbr.":"imbramowicki",
                     "jędrz.":"jędrzejowski",
                     "kal.":"kaliski",
                     "Klar.":"Klaryski",
                     "koprz.":"koprzywnicki",
                     "ksiąs.":"ksiąski",
                     "lub.":"lubelski",
                     "łęcz.":"łęczycki",
                     "magd.":"magdeburski",
                     "maz.":"mazowiecki",
                     "miech.":"miechowski",
                     "mog.":"mogilski",
                     "mstow.":"mstowski",
                     "niem.":"niemiecki",
                     "opocz.":"opoczyński",
                     "ośw.":"oświęcimski",
                     "pilzn.":"pilzneński",
                     "pol.":"polski",
                     "poł.":"połaniecki",
                     "pozn.":"poznański",
                     "pras.":"praskie",
                     "prosz.":"proszowski",
                     "przem.":"przemyski",
                     "rad.":"radomski",
                     "roz.":"rozpierski",
                     "san.":"sanocki",
                     "sand.":"sandomierski",
                     "sądec.":"sądecki",
                     "sieciech.":"sieciechowski",
                     "sier.":"sieradzki",
                     "siew.":"siewierski",
                     "Siew.":"siewierski",
                     "staniąt.":"staniątecki",
                     "starosądec.":"starosądecki",
                     "szczyrz.":"szczyrzycki",
                     "śl.":"śląski",
                     "średz.":"średzki",
                     "świętokrz.":"świętokrzyski",
                     "tyn.":"tyniecki",
                     "wąch.":"wąchocki",
                     "węg.":"węgierski",
                     "wiel.":"wieluński",
                     "wiśl.":"wiślicki",
                     "wojn.":"wojnicki",
                     "zator.":"Zatorski",
                     "zawich.":"zawichojski",
                     "zwierzyn.":"zwierzyniecki",
                     "żarn.":"żarnowski",
                     "żyd.":"żydowski"
                    }
    elif slownik == "Czersk":
        shortcuts = {"bełs.":"bełski",
                     "bial.":"bialski",
                     "biel.":"bielski",
                     "bł.":"błoński",
                     "brzeskuj.":"brzesko-kujawskie",
                     "chełm.":"chełmiński",
                     "ciech.":"ciechanowski",
                     "czer.":"czerski",
                     "czerw.":"czerwiński",
                     "droh.":"drohicki",
                     "garw.":"garwoliński",
                     "gąb.":"gąbiński",
                     "gnieźn.":"gnieźnieński",
                     "gost.":"gostyniński",
                     "grodz.":"grodziecki",
                     "grój.":"grójecki",
                     "kal.":"kaliski",
                     "kam.":"kamieniecki",
                     "koln.":"kolneński",
                     "krak.":"krakowski",
                     "krzyż.":"krzyżacki",
                     "lit.":"litewski",
                     "liw.":"liwski",
                     "lub.":"lubelski",
                     "łęcz.":"łęczycki",
                     "łomż.":"łomżyński",
                     "łuk.":"łukowski",
                     "magd.":"magdeburski",
                     "mak.":"makowski",
                     "maz.":"mazowiecki",
                     "mław.":"mławski",
                     "nowom.":"nowomiejski",
                     "ostroł.":"ostrołęcki",
                     "ostrow.":"ostrowski",
                     "płoc.":"płocki",
                     "płoń.":"płoński",
                     "pol.":"polski",
                     "pozn.":"poznański",
                     "przas.":"przasnyski",
                     "pułt.":"pułtuski",
                     "rac.":"raciąski",
                     "rad.":"radomski",
                     "radz.":"radziłowski",
                     "raw.":"rawski",
                     "roż.":"rożański",
                     "sand.":"sandomierski",
                     "sąch.":"sąchocki",
                     "ser.":"serocki",
                     "sierp.":"sierpecki",
                     "soch.":"sochaczewski",
                     "stęż.":"stężycki",
                     "szreń.":"szreński",
                     "tarcz.":"tarczyński",
                     "war.":"warecki",
                     "warsz.":"warszawski",
                     "wąs.":"wąsoski",
                     "węg.":"węgierski",
                     "wil.":"wileński",
                     "wis.":"wiski",
                     "wlkp.":"wielkopolski",
                     "włocł.":"włocławski",
                     "wysz.":"wyszogrodzki",
                     "zakr.":"zakroczymski",
                     "zawkrz.":"zawkrzeński",
                     "ziem.":"ziemski"
                    }
    elif slownik == 'Lublin':
        shortcuts = {"krak.":"krakowski",
                      "lub.":"lubelski",
                      "łuk.":"łukowski",
                      "magd.":"magdeburski",
                      "rad.":"radomski",
                      "sand.":"sandomierski",
                      "węg.":"węgierski"
                    }
    elif slownik == 'Lublin_zaginione':
        shortcuts = [] # nie ma skrótów geograficznych?
    elif slownik == 'Płock':
        shortcuts = {"biel.":"bielski",
                     "bł.":"błoński",
                     "chełm.":"chełmiński",
                     "ciech.":"ciechanowski",
                     "czer.":"czerski",
                     "czerw.":"czerwiński",
                     "dobrz.":"dobrzyński",
                     "flam.":"flamandzki",
                     "gniezn.":"gnieźnieński",
                     "gost.":"gostyński",
                     "grodz.":"grodzki",
                     "grój.":"grójecki",
                     "kam.":"kamieniecki",
                     "kol.":"koleński",
                     "krak.":"krakowski",
                     "krzyż.":"krzyżacki",
                     "kuj.":"kujawski",
                     "liw.":"liwski",
                     "łęcz.":"łęczycki",
                     "łom.":"łomżyński",
                     "łow.":"łowicki",
                     "magd.":"magdeburski",
                     "mak.":"makowski",
                     "maz.":"mazowiecki",
                     "miej.":"miejski",
                     "mław.":"mławski",
                     "niedzb.":"niedzborski",
                     "now.":"nowomiejski",
                     "nowogr.":"nowogrodzki",
                     "ostroł.":"ostrołęcki",
                     "ostrow.":"ostrowski",
                     "płoc.":"płocki",
                     "płoń.":"płoński",
                     "pom.":"pomorski",
                     "pozn.":"poznański",
                     "pras.":"praski",
                     "przas.":"przasnyski",
                     "pułt.":"pułtuski",
                     "rac.":"raciąski",
                     "raw.":"rawski",
                     "róż.":"różański",
                     "sąch.":"sąchocki",
                     "ser.":"serocki",
                     "siel.":"sieluński",
                     "sierp.":"sierpecki",
                     "soch.":"sochaczewski",
                     "szr.":"szreński",
                     "śląs.":"śląski",
                     "war.":"warecki",
                     "warsz.":"warszawski",
                     "wąs.":"wąsoski",
                     "węg.":"węgierski",
                     "wis.":"wiski",
                     "wysz.":"wyszogrodzki",
                     "zakr.":"zakroczymski",
                     "zamb.":"zambrowski",
                     "zawkrz.":"zawkrzeński"
                     }
    elif slownik == 'Podlasie':
        shortcuts = {
                    "august.":"augustowski",
                    "bełs.":"bełski",
                    "bial.":"bialski",
                    "biel.":"bielski",
                    "bł.":"błoński",
                    "brań.":"brański",
                    "brzeskuj":"brzesko-kujawskie",
                    "brzeslit.":"brzesko-litewskie",
                    "chełm.":"chełmiński",
                    "ciech.":"ciechanowski",
                    "czer.":"czerski",
                    "droh.":"drohicki",
                    "garw.":"garwoliński",
                    "gąb.":"gąbiński",
                    "gnieźn.":"gnieźnieński",
                    "goniądz.":"goniądzki",
                    "gost.":"gostyniński",
                    "grodz.":"grodzki",
                    "grój.":"grójecki",
                    "kam.":"kamieniecki",
                    "knysz.":"knyszyński",
                    "koln.":"kolneński",
                    "krak.":"krakowski",
                    "lit.":"litewski",
                    "liw.":"liwski",
                    "łęcz.":"łęczycki",
                    "łomż.":"łomżyński",
                    "łuk.":"łukowski",
                    "mak.":"makowski",
                    "maz.":"mazowiecki",
                    "mieln.":"mielnicki",
                    "mław.":"mławski",
                    "mszczon.":"mszczonowski",
                    "niedz.":"niedzborski",
                    "nowom.":"nowomiejski",
                    "nur.":"nurski",
                    "ostroł.":"ostrołęcki",
                    "ostrow.":"ostrowski",
                    "płoc.":"płocki",
                    "płoń.":"płoński",
                    "podl.":"podlaski",
                    "pozn.":"poznański",
                    "przas.":"przasnyski",
                    "pułt.":"pułtuski",
                    "rac.":"raciąski",
                    "radz.":"radziłowski",
                    "rajgr.":"rajgrodzki",
                    "raw.":"rawski",
                    "roż.":"rożański",
                    "sand.":"sandomierski",
                    "sąch.":"sąchocki",
                    "ser.":"serocki",
                    "sierp.":"sierpecki",
                    "soch.":"sochaczewski",
                    "sur.":"suraski",
                    "szreń.":"szreński",
                    "tarcz.":"tarczyński",
                    "tykoc.":"tykociński",
                    "war.":"warecki",
                    "wąs.":"wąsoski",
                    "węg.":"węgierski",
                    "wil.":"wileński",
                    "wis.":"wiski",
                    "włocł.":"włocławski",
                    "włodzim.":"włodzimierski",
                    "wysz.":"wyszogrodzki",
                    "zakr.":"zakroczymski",
                    "zamb.":"zambrowski",
                    "zawkrz.":"zawkrzeński"
        }
    elif slownik == 'Poznań':
        shortcuts = {"bab.":"babimojski",
                     "biech.":"biechowski",
                     "bledz.":"bledzewski",
                     "bnin.":"bniński",
                     "brand.":"brandenburski",
                     "buk.":"bukowski",
                     "bydg.":"bydgoski",
                     "chełm.":"chełmiński",
                     "czarnk.":"czarnkowski",
                     "dobrz.":"dobrzyński",
                     "flam.":"flamandzki",
                     "frank.":"frankoński",
                     "giec.":"giecki",
                     "głog.":"głogowski",
                     "gnieźn.":"gnieźnieński",
                     "inowrocł.":"inowrocławski",
                     "kal.":"kaliski",
                     "kam.":"kamieński",
                     "karczm.":"karczmarski",
                     "karz.":"karzecki",
                     "kcyn.":"kcyński",
                     "klar.":"klaryski",
                     "kon.":"koniński",
                     "kostrz.":"kostrzyński",
                     "kośc.":"kościański",
                     "krak.":"krakowski",
                     "krob.":"krobski",
                     "krusz.":"kruszwicki",
                     "krzyw.":"krzywiński",
                     "ksiąs.":"ksiąski",
                     "kuj.":"kujawski",
                     "lądz.":"lądzki",
                     "lub.":"lubiński",
                     "łekn.":"łekneński",
                     "łęcz.":"łęczycki",
                     "magd.":"magdeburski",
                     "maz.":"mazowiecki",
                     "miej.":"miejski",
                     "międz.":"międzyrzecki",
                     "młp.":"małopolski",
                     "mod.":"moderski",
                     "mogil.":"mogileński",
                     "nak.":"nakielski",
                     "niem.":"niemiecki",
                     "ober.":"oberski",
                     "odol.":"odolanowski",
                     "ołob.":"ołobocki",
                     "opat.":"opatówecki",
                     "ostrz.":"ostrzeszowski",
                     "owin.":"owiński",
                     "parad.":"paradyski",
                     "pobiedz.":"pobiedziski",
                     "pol.":"polski",
                     "pom.":"pomorski",
                     "pozn.":"poznański",
                     "pras.":"praski",
                     "przem.":"przemęcki",
                     "pszcz.":"pszczewski",
                     "pyzdr.":"pyzdrski",
                     "radz.":"radzimski",
                     "rogoz.":"rogoziński",
                     "sant.":"santocki",
                     "sier.":"sieradzki",
                     "soł.":"sołecki",
                     "starogr.":"starogrodzki",
                     "śląs.":"śląski",
                     "średz.":"średzki",
                     "śrem.":"śremski",
                     "trzem.":"trzemeszeński",
                     "wał.":"wałecki",
                     "wągr.":"wągrowiecki",
                     "węg.":"węgierski",
                     "wielich.":"wielichowski",
                     "wlkp.":"wielkopolski",
                     "włocł.":"włocławski",
                     "wrocł.":"wrocławski",
                     "wsch.":"wschowski",
                     "zbąsz.":"zbąszyński",
                     "żnin.":"żniński",
                     "żon.":"żoński"
                     }
    elif slownik == 'Sanok':
        shortcuts = {"biec.":"biecki",
                     "gr.":"grodzki",
                     "krak.":"krakowski",
                     "krośn.":"krośnieński",
                     "lw.":"lwowski",
                     "łac.":"łaciński",
                     "magd.":"magdeburski",
                     "niem.":"niemiecki",
                     "pilzn.":"pilzneński",
                     "pol.":"polski",
                     "przem.":"przemyski",
                     "samb.":"samborski",
                     "san.":"sanocki",
                     "sandom.":"sandomierski",
                     "sądec.":"sądecki",
                     "węg.":"węgierski",
                     "żydacz.":"żydaczowski"
                     }
    elif slownik == 'Wieluń':
        shortcuts = {"bolesł.":"bolesławiecki",
                     "gnieźń.":"gnieźnieński",
                     "kal.":"kaliski",
                     "krak.":"krakowski",
                     "magd.":"magdeburski",
                     "niem.":"niemiecki",
                     "ostrz.":"ostrzeszowski",
                     "Ostrz.":"ostrzeszowski",
                     "pozn.":"poznański",
                     "sier.":"sieradzki",
                     "średz.":"średzki",
                     "Vel.":"Velunensis",
                     "Viel.":"Velunensis",
                     "Vyel.":"Velunensis",
                     "wiel.":"wieluński",
                     "wrocł.":"wrocławski"
                     }
    elif slownik == 'Wyszogród':
        shortcuts = {"biel":"bielski",
                     "Bł.":"błoński",
                     "chełm.":"chełmiński",
                     "Ciech.":"ciechanowski",
                     "czer.":"czerski",
                     "Czerw.":"czerwiński",
                     "dobrz.":"dobrzyński",
                     "gniezn.":"gnieźnieński",
                     "Gost.":"gostyński",
                     "gost.":"gostyński",
                     "grój.":"grójecki",
                     "Kam.":"kamieniecki",
                     "kol.":"koloński",
                     "liw.":"liwski",
                     "Łęcz.":"łęczycki",
                     "łom.":"łomżyński",
                     "Łow.":"łowicki",
                     "łow.":"łowicki",
                     "mak.":"makowski",
                     "Maz.":"mazowiecki",
                     "mław.":"mławski",
                     "Niedz.":"niedzborski",
                     "nur.":"nurski",
                     "Ostroł.":"ostrołęcki",
                     "ostrow.":"ostrowski",
                     "płoc.":"płocki",
                     "płoń.":"płoński",
                     "pozn.":"poznański",
                     "pras.":"praski",
                     "Przas.":"przasnyski",
                     "przas.":"przasnyski",
                     "Pułt.":"pułtuski",
                     "pułt.":"pułtuski",
                     "Rac.":"raciąski",
                     "rac.":"raciąski",
                     "raw.":"rawski",
                     "róż.":"różański",
                     "Ser.":"serocki",
                     "siel.":"sieluński",
                     "Soch.":"sochaczewski",
                     "soch.":"sochaczewski",
                     "Szr.":"szreński",
                     "war.":"warecki",
                     "Warsz.":"warszawski",
                     "warsz.":"warszawski",
                     "Wąs.":"wąsoski",
                     "węg.":"węgierski",
                     "wis.":"wiski",
                     "Wysz.":"wyszogrodzki",
                     "wysz.":"wyszogrodzki",
                     "Zakr.":"zakroczymski",
                     "zakr.":"zakroczymski",
                     "Zamb.":"zambrowski",
                     "zamb.":"zambrowski",
                     "zawkrz.":"zawkrzeński"
                     }
    elif slownik == 'Warszawa':
        shortcuts = {"bełs.":"bełski",
                     "bł.":"błoński",
                     "chełm.":"chełmiński",
                     "ciech.":"ciechanowski",
                     "czer.":"czerski",
                     "czerw.":"czerwiński",
                     "gnieźn.":"gnieźnieński",
                     "gost.":"gostyniński",
                     "grój.":"grójecki",
                     "kam.":"kamieniecki",
                     "krak.":"krakowski",
                     "krzyż.":"krzyżacki",
                     "liw.":"liwski",
                     "łęcz.":"łęczycki",
                     "łomż.":"łomżyński",
                     "magd.":"magdeburski",
                     "mak.":"makowski",
                     "maz.":"mazowiecki",
                     "nur.":"nurski",
                     "ostroł.":"ostrołęcki",
                     "ostrow.":"ostrowski",
                     "płoc.":"płocki",
                     "płoń.":"płoński",
                     "pol.":"polski",
                     "pozn.":"poznański",
                     "pras.":"praski",
                     "pułt.":"pułtuski",
                     "rac.":"raciąski",
                     "raw.":"rawski",
                     "roż.":"rożański",
                     "sand.":"sandomierski",
                     "ser.":"serocki",
                     "soch.":"sochaczewski",
                     "tarcz.":"tarczyński",
                     "war.":"warecki",
                     "warsz.":"warszawski",
                     "węg.":"węgierski",
                     "wis.":"wiski",
                     "wlkp.":"wielkopolski",
                     "włocł.":"włocławski",
                     "wysz.":"wyszogrodzki",
                     "zakr.":"zakroczymski",
                     "zawkrz.":"zawkrzeński"
                     }
    elif slownik == 'Liw':
        shortcuts = {"bełs.":"bełski",
                     "bial.":"bialski",
                     "biel.":"bielski",
                     "bł.":"błoński",
                     "chełm.":"chełmiński",
                     "ciech.":"ciechanowski",
                     "czer.":"czerski",
                     "czerw.":"czerwiński",
                     "droh.":"drohicki",
                     "garw.":"garwoliński",
                     "gąb.":"gąbiński",
                     "gnieźn.":"gnieźnieński",
                     "gost.":"gostyniński",
                     "grodz.":"grodzki",
                     "grój.":"grójecki",
                     "kam.":"kamieniecki",
                     "koln.":"kolneński",
                     "krak.":"krakowski",
                     "król.":"królewski",
                     "krzyż.":"krzyżacki",
                     "lit.":"litewski",
                     "liw.":"liwski",
                     "łęcz.":"łęczycki",
                     "łomż.":"łomżyński",
                     "łuk.":"łukowski",
                     "magd.":"magdeburski",
                     "mak.":"makowski",
                     "maz.":"mazowiecki",
                     "mław.":"mławski",
                     "mszczon.":"mszczonowski",
                     "niedz.":"niedzborski",
                     "niem.":"niemiecki",
                     "nowom.":"nowomiejski",
                     "nur.":"nurski",
                     "ostroł.":"ostrołęcki",
                     "ostrow.":"ostrowski",
                     "płoc.":"płocki",
                     "płoń.":"płoński",
                     "pol.":"polski",
                     "pozn.":"poznański",
                     "pras.":"praski",
                     "przas.":"przasnyski",
                     "pułt.":"pułtuski",
                     "rac.":"raciąski",
                     "radz.":"radziłowski",
                     "raw.":"rawski",
                     "roż.":"rożański",
                     "sand.":"sandomierski",
                     "sąch.":"sąchocki",
                     "ser.":"serocki",
                     "sierp.":"sierpecki",
                     "soch.":"sochaczewski",
                     "szreń.":"szreński",
                     "tarcz.":"tarczyński",
                     "war.":"warecki",
                     "warsz.":"warszawski",
                     "wąs.":"wąsoski",
                     "węg.":"węgierski",
                     "wil.":"wileński",
                     "wis.":"wiski",
                     "wlkp.":"wielkopolski",
                     "włocł.":"włocławski",
                     "wysz.":"wyszogrodzki",
                     "zakr.":"zakroczymski",
                     "zamb.":"zambrowski",
                     "zawkrz.":"zawkrzeński",
                     "ziem.":"ziemski"
                     }
    elif slownik == 'Czersk':
        shortcuts = {"bełs.":"bełski",
                     "bial.":"bialski",
                     "biel.":"bielski",
                     "bł.":"błoński",
                     "brzeskuj.":"brzesko-kujawskie",
                     "chełm.":"chełmiński",
                     "ciech.":"ciechanowski",
                     "czer.":"czerski",
                     "czerw.":"czerwiński",
                     "droh.":"drohicki",
                     "garw.":"garwoliński",
                     "gąb.":"gąbiński",
                     "gnieźn.":"gnieźnieński",
                     "gost.":"gostyniński",
                     "grodz.":"grodziecki",
                     "grój.":"grójecki",
                     "kal.":"kaliski",
                     "kam.":"kamieniecki",
                     "koln.":"kolneński",
                     "krak.":"krakowski",
                     "krzyż.":"krzyżacki",
                     "lit.":"litewski",
                     "liw.":"liwski",
                     "lub.":"lubelski",
                     "łęcz.":"łęczycki",
                     "łomż.":"łomżyński",
                     "łuk.":"łukowski",
                     "magd.":"magdeburski",
                     "mak.":"makowski",
                     "maz.":"mazowiecki",
                     "mław.":"mławski",
                     "młyń.":"młyński",
                     "mszczon.":"mszczonowski",
                     "niem.":"niemiecki",
                     "nur.":"nurski",
                     "ostroł.":"ostrołęcki",
                     "ostrow.":"ostrowski",
                     "płoc.":"płocki",
                     "płoń.":"płoński",
                     "pozn.":"poznański",
                     "przas.":"przasnyski",
                     "pułt.":"pułtuski",
                     "rac.":"raciąski",
                     "rad.":"radomski",
                     "radz.":"radziłowski",
                     "raw.":"rawski",
                     "roż.":"rożański",
                     "sand.":"sandomierski",
                     "sąch.":"sąchocki",
                     "ser.":"serocki",
                     "sierp.":"sierpecki",
                     "soch.":"sochaczewski",
                     "stęż.":"stężycki",
                     "szreń.":"szreński",
                     "tarcz.":"tarczyński",
                     "war.":"warecki",
                     "warsz.":"warszawski",
                     "wąs.":"wąsoski",
                     "wil.":"wileński",
                     "wlkp.":"wielkopolski",
                     "włocł.":"włocławski",
                     "wysz.":"wyszogrodzki",
                     "zakr.":"zakroczymski",
                     "zawkrz.":"zawkrzeński"
                    }

    return shortcuts


def expand_abbr(value: str, typ: list) -> str:
    """ funkcja rozwija skrót jeżeli jest"""
    tmp = ''
    for t in typ:
        if t in value:
            tmp = value.replace(t, '').strip()
            break

    if tmp and tmp in skroty:
        value = skroty[tmp]
    else:
        value = tmp

    return value


def clear_value(value: str, abbr_list: list) -> str:
    """ funkcja usuwa podany skrót z tekstu """

    for abbr in abbr_list:
        if abbr in value:
            value = value.replace(abbr, '').strip()
        else:
            abbr = abbr[0].upper() + abbr[1:]
            value = value.replace(abbr, '').strip()

    return value.strip()

def czy_miejscowosc(value_header: str, value_p0: str) -> bool:
    """ sprawdza czy hasło jest miejscowością """
    result = True

    words_header = ['tenuta', 'okręg', 'kasztelania', 'księstwo', 'starostwo', 'lacus', 'palus',
             'lutum', 'collum', 'powiat', 'ziemia', 'districtus', 'oficjalat', 'archidiakonat',
             'diecezja', 'województwo', 'klucz', 'klasztor', 'opactwo', 'łany', 'dystrykt', 'dekanat',
             'dobra', 'przekazy', 'niezrealizowana', 'SUMMUM']

    # słowa które wykluczają miejscowość (początek punktu 0)
    words = ['→', 'kasztelania', 'księstwo', 'starostwo', 'districtus', 'oficjalat', 'klucz dóbr',
             'lutum', 'collum', 'leśnictwo', 'tenuta', 'polana', 'ziemia', 'klucz',
             'diecezja', 'struga', 'jezioro', 'jeziorko', 'łąka', 'rzeka', 'las',
             'staw', 'puszcza', 'bagna', 'bagno', 'góra', 'trakt', 'rzeka',
             'dopływ', 'fossula', 'bagnisko', 'droga', 'potok', 'pole',
             'bór', 'miejsce', 'zarośla', 'pola', 'rz.', 'błonia', 'pastwiska',
             'pastwisko', 'ogrody', 'lasem', 'źródło', 'strumień', 'padół',
             'rzeczka', 'gaj', 'łąki','lasy', 'ługi', 'uroczysko', 'grodzisko',
             'źródło', 'obiekt', 'góra', 'góra?', 'skały', 'wąwóz', 'wzgórze',
             'zamek', 'castrum', 'kuźnica', 'wzniesienie', 'kopiec', 'dolina',
             'pasmo', 'krzaki', 'gaje', 'gaj', 'rów', 'teren', 'role', 'lasów', 'lasków',
             'huty', 'potok?', 'wodończa?', 'dziedzina', 'zagłębienie', 'potoki',
             'KARCZMY', 'dolina?', 'potok?', 'dąbrowa', 'mosty', 'opactwo',
             'karczma', 'wierzchołek', 'dąbrowa', 'merica', 'archidiakonat', 'błoto',
             'woda', 'grzbiet górski', 'grzbiet', 'mostek', 'jez.', 'łożysko', 'stawek',
             'padół', 'pola', 'pola?', 'ulica', 'most', 'debrz', 'las?', 'starorzecze',
             'słup', 'staw-odroślisko', 'odroślisko', 'mostki', 'grupa stawów', 'stawów',
             'góry', 'stawy', 'smug', 'potok', 'potoków', 'górka', 'skała', 'lasek',
             'wąwóz', 'wąwozem', 'młynówka', 'rola', 'stawów', 'nazwa brzegu Wisły', 'brzegu',
             'grobla', 'miejsce', 'wyspa', 'kępa'
             ]

    # czy pomijać? : ['młyn', 'folwark', 'karczma', 'karczmy', 'kuźnice', 'kuźnica', 'włość']

    legalne = ['wieś', 'osada', 'grodzisko', 'gród', 'wioska', 'przysiółek',
               'miasto', 'przedmieście']

    if value_header.endswith(';'):
        value_header = value_header[:-1]
    if value_header.endswith('.'):
        value_header = value_header[:-1]

    result = all_caps(value_header)

    # przeszukiwanie nazwy
    if result:
        for word in words_header:
            #lista_value_header = value_header.split()
            lista_value_header = re.split(',|-| |―', value_header)
            for word_h in lista_value_header:
                if word.lower() == word_h.lower():
                    result = False
                    break
            if not result:
                break

    # przeszukanie treści pierwszego punktu lub akapitu
    if result and value_p0:
        # usunięcie wyrażeń mylących algorytm, że to nie jest miejscowość
        wyrazenia = ['nad rz.', 'nad rzeką', 'nad rzeczką', 'u ujścia rz.',
                     'na lewym brzegu rz.', 'na prawym brzegu rz.']
        for wyr in wyrazenia:
            if wyr in value_p0:
                value_p0 = double_space(value_p0.replace(wyr, ''))

        # usuwanie nawiasu z odmiankami jeżeli jest
        if value_p0.startswith('('):
            pos = value_p0.find(')')
            if pos != -1:
                value_p0 = value_p0[pos:].strip()

        lista = value_p0.split(' ')
        lista = [l.replace(',','').replace(';','').strip() for l in lista]
        lista = [l for l in lista if len(l) > 2]
        if lista:
            lista = lista[:20]

        for element in lista:
            if element.lower() in words:
                result = False
                break

    return result


def parafia_nazwa(value: str, nazwa_miejscowosci: str) -> str:
    """ Weryfikuje i poprawia nazwę parafii, jeżeli jest tylko 1 duża litera,
        co zwykle oznacza parafię zgodną z nazwą miejscowości, oraz eliminacja
        przypadkowych artefaktów typu par. 110
    """
    result = value

    value_a = ''
    if value.endswith('.'):
        value_a = value[:-1]
    elif len(value) == 1 and value.isupper():
        value_a = value

    if value_a:
        if nazwa_miejscowosci.lower().startswith(value_a.lower()):
            result = nazwa_miejscowosci
    else:
        if value.isdigit():
            result = ''

    return result


def unique(sequence):
    """ zwraca unikalną listę """
    seen = set()
    return [x for x in sequence if not (x in seen or seen.add(x))]


def get_odmianki_standard():
    """ odmianki standard """
    odmianki = ''

    pattern = r'\(.*?\)'
    match = re.search(pattern=pattern, string=p0)
    if match:
        odmianki_txt = match.group()
        odmianki_txt = odmianki_txt.replace('(', '').replace(')', '')

        pomijane = ['Kraków'] # w tym słowniku jets inny układ w odmiankach
        if slowniki[nr] not in pomijane:
            pos = odmianki_txt.find(';')
            if pos != -1:
                # czy to sam rok oddzielony ; ? po takim roku czasem są odmianki
                przed = odmianki_txt[:pos].strip()
                if przed.isdigit():
                    odmianki_txt = odmianki_txt[pos+1:]
                    pos = odmianki_txt.find(';')
                    if pos != -1:
                        odmianki_txt = odmianki_txt[:pos]
                else:
                    odmianki_txt = odmianki_txt[:pos]

        odmianki_lista = odmianki_txt.split(',')
        for item in odmianki_lista:
            # czyszczenie
            words_to_strip = ['[!]', 'w źr. także', 'w źr zw.',
                                'poł. XV w.', 'w źr. zw. też', ' ok.']
            for word_s in words_to_strip:
                item = item.replace(word_s, '').strip()
            item = double_space(item).strip()

            if nr in [3]:
                pauzy = [' – ', ' — ', ' —','—', ' - ']
                for pauza in pauzy:
                    pos_pauza = item.find(pauza)
                    if pos_pauza != -1:
                        item = item[:pos_pauza]

            # czy element zawiera słowa przypominające nazwę własną?
            pattern_nazwa = r'[A-ZĄŚĘŻŹĆŃŁÓ]{1}[\w]+'
            match_nazwa = re.search(pattern_nazwa, item)
            if not match_nazwa:
                continue

            # if nazwa == 'RZEPINEK':
            #     print()

            pattern_year = r'\[?[a\. ]{0,}\d{4}(-\d{1,4})?\s?[n\.]{0,}\]?'
            #match_year = re.search(pattern=pattern_year, string=item)
            match_years = [x.group() for x in re.finditer(pattern_year, item)]
            if match_years:
                odmianka = item
                for year in match_years:
                    odmianka = odmianka.replace(year, '').strip()
            else:
                year = ''
                odmianka = item.strip()

            if odmianka and ']' in odmianka:
                odmianka = odmianka.replace(']', '')
            if odmianka and '[' in odmianka:
                odmianka = odmianka.replace('[', '')

            pattern_to_strip = [r'w\s+źr\.\s+z\s+[XIV-]+\s+w\.\s+także',
                                r'w\s+[XVI-]+\s+w\.\s+także',
                                r'w\s+[XVI-]+\s+w\.\s+też',
                                r'w\s+źr\.\s+do\s+poł\.\s+[XVI-]+\s+w\.',
                                r'od\s+końca\s+[XVI-]+\s+w\.',
                                r'koniec\s+[XVI-]+\s+w\.',
                                r'w\s+źr\.\s+',
                                r'[XVI-]+\s+w\.']
            for pattern in pattern_to_strip:
                odmianka = re.sub(pattern, '', odmianka).strip()
            odmianka = double_space(odmianka).strip()


            # wskazanie na odnośnik to chyba nie odmianka?
            if odmianka.startswith('→'):
                odmianka = ''

            if len(item)< 2:
                continue

            # czy to na pewno nie referencja do publikacji?
            pattern = r'\s{1}\d{2,3}-?'
            match_ref = re.search(pattern, odmianka)
            pattern = r'\s{1}\d{2,3}-\d{1}'
            match_ref2 = re.search(pattern, odmianka)
            if match_ref or match_ref2:
                continue

            # dalsze czyszczenie
            odmianka = re.sub(r'\d{1,3}', '', odmianka)

            if nr in [3, 7]: # Kraków, Poznań
                clear_words = [';', '- KK', '- SP', '- KUJ', '- KSN', ' - MS', ' - AGZ',
                            '- Księga ławnicza', '- ZDM', ' ze wzm.', ' z kopii z',
                            ' - ZDK', 'Nieznany dokument →p.', '- reg. ZDM', '-Pol.',
                            '! rps XV w. ', 'rps XV w. ', 'or. i trans.', 'interpol.'
                            '- W. Smolarkiewicz', '- Sroka Dokumenty', 'fals.',
                            '- M. Zieleniewski', 'Illustrowany opis c. k. zakładu zdrojowego w Krynicy',
                            ' z or.', '- Mp.', 'z fals.', 'z transumptu zr.', ' rps ',
                            '- Pol.', '- Wp.', ' z kopii,', ' kop.', ' z kopii ',
                            'z kopii', ' s.', 'w transumpcie z', 'Wieś →p.', 'Parafia → p.',
                            's. XXXVI', 'fals. z XIVw.', 'Mp.', ' z pocz.,', 'AGZ',
                            'wzm. z DLb.', 'DLb.', '→ niżej', 'naśladującej or. z i jej odpisów',
                            ' zapewne z', 'z XV i', '→ DH przed', '― fals.', 'z dok. interpol.',
                            'wzmianka zr.', 'między ', 'z MPH', 'z pocz.', 'poł.',
                            ' z transumptu z r', ' fals. z', 'forma zmodernizowana', 'kop. ',
                            ' w or. z', ' w fals.', ' z reg. z r', 'z dok. interpolowanego po rznanego',
                            ' fals. z', 'wg ind. WAP', 'SP', ' z reg. z', 'w zestawieniu ',
                            'interpol.? kop', 'rkps ior.', 'zw. też', 'ikop ', 'rps ', 'fals ',
                            ' -przekaz z', ' z zr.', '→ ', ' z DHn.', 'trans.', 'or. i trans', ' or.',
                            'schyłek '
                            ]
                for cl in clear_words:
                    odmianka = odmianka.replace(cl, '').strip()

            odmianka = double_space(odmianka).strip()
            if odmianka.endswith(' z') or odmianka.endswith(' w') or odmianka.endswith(' p.'):
                odmianka = odmianka[:-2].strip()
            if odmianka.startswith('or.'):
                odmianka = odmianka[3:].strip()
            if odmianka.startswith('kop') and odmianka[3].isupper():
                odmianka = odmianka[3:].strip()
            if odmianka.startswith('ok') and (odmianka[2].isupper() or odmianka[2] == '„'):
                odmianka = odmianka[2:].strip()
            if odmianka.startswith('obl') and odmianka[3].isupper():
                odmianka = odmianka[3:].strip()
            if odmianka.startswith('! kop') and odmianka[5].isupper():
                odmianka = odmianka[5:].strip()
            if odmianka.startswith('i kop') and odmianka[5].isupper():
                odmianka = odmianka[5:].strip()
            if odmianka.startswith('przed') and odmianka[5].isupper():
                odmianka = odmianka[5:].strip()
            if odmianka.startswith('reg.'):
                odmianka = odmianka[4:].strip()
            odmianka = odmianka.replace(r'/','')
            if odmianka.startswith('reg') and odmianka[3].isupper():
                odmianka = odmianka[3:].strip()
            if odmianka.startswith('trans') and odmianka[5].isupper():
                odmianka = odmianka[5:].strip()
            if odmianka.startswith('fals') and odmianka[4].isupper():
                odmianka = odmianka[4:].strip()
            if odmianka.strip() == 'DK':
                odmianka = ''
            if odmianka.strip() == 'PrU':
                odmianka = ''
            if odmianka.strip() == 'Pr U':
                odmianka = ''
            if odmianka.strip() == 'BK':
                odmianka = ''
            if odmianka.strip() == 'schyłek':
                odmianka = ''

            if odmianka:
                odmianki += odmianka + ', '

    if odmianki.endswith(', '):
        odmianki = odmianki[:-2]

    return odmianki

def get_odmianki_poznan():
    """ get odmianki Poznan """
    odmianki = ''

    if '[Pow. ' in p0:
        pos = p0.find('[Pow. ')
        p_odmianki = p0[:pos]
    else:
        p_odmianki = p0

    o_list = xsplit(p_odmianki)
    odm_list = []
    for o_item in o_list:
        o_item = o_item.strip()
        if o_item[0].isdigit():
            pos = o_item.find('(')
            if pos != -1:
                o_item = o_item[:pos].strip()

            o_item = re.sub(r'\[.*?\]', '', o_item)
            o_item = re.sub(r'^\d{4}[\?]?', '', o_item)
            # czyszczenie
            c_words = ['kop. ', 'or. ', 'XV w.', 'XIV w.', 'XIII w.', 'XII w.', 'XVI w.',
                       'XX w.', 'XIX w.', 'XVII w.', 'XVIII w.',
                       '[fals. ok. ]', '[fals. z końca ]', 'regesty z ', '[fals. ] ',
                       '[XVII-XVIII w.]', '[]', 'obl. ', 'fals. koniec',
                       'XVII/XVIII w.', 'XV-XVII w.', 'XIX w.', 'wzm. z ', 'wzm.' , 'Wp. nr',
                       'rpsy XIV', 'rkpsy z XIV', 'wg relacji', 'XIV', 'XVIII', 'XVII', 'XVI',
                       'XIX', 'XIII', 'XII', 'XI', 'z 2. poł.', '2. poł.', 'z 1. poł.', '1. poł.']
            for c_w in c_words:
                o_item = o_item.replace(c_w, '')
            o_item = re.sub(r'\d{4}-\d{2}', '', o_item).strip()
            o_item = re.sub(r'\d{4}', '', o_item).strip()

            patterns = [r'^\d{1,2}\s+km\s+', r'^\d{1,2}\s?,\s?\d{1}\s+km\s+']
            for pattern in patterns:
                match = re.search(pattern, o_item)
                if match:
                    o_item =''
                    break

            # jeżeli same cyfry
            if o_item.isdigit():
                o_item =''

            if o_item:
                odm_list.append(o_item.strip())
        else:
            continue

    odmianki = ', '.join(odm_list)

    return odmianki


def xsplit(s:str):
    """ funkcja dzieli przekazany tekst na wiersze według przecinka pomijając
        jednak przecinki w nawiasach, zwraca listę
    """
    parts = []
    bracket_level = 0
    current = []
    # trick to remove special-case of trailing chars
    for c in (s + ","):
        if c == "," and bracket_level == 0:
            parts.append("".join(current))
            current = []
        else:
            if c == "(":
                bracket_level += 1
            elif c == ")":
                bracket_level -= 1
            current.append(c)
    return parts


# -------------- Główny program ------------------------------------------------
if __name__ == '__main__':

    for nr in range(4, 5):
        print(f"Słownik: {slowniki[nr]} ({nr})")
        skroty = shortcut(slowniki[nr])

        output_path = Path('.').parent / f'shg_ahp/shg_ahp_{slowniki[nr]}_new.xlsx'
        input_path = Path('.').parent / f'output/hasla_{nr}_pkt.csv'
        odnosniki_path = Path('.').parent / f'output/odnosniki_{nr}.txt'
        odnosniki_path_new = Path('.').parent / f'output/odnosniki_{nr}_new.txt'

        if os.path.exists(odnosniki_path_new):
            os.remove(odnosniki_path_new)

        wb = openpyxl.Workbook()
        ws = wb.active

        # nagłówek, tylko podział wg akapitów
        point_sheet_columns = ['SHG_ID', 'Nazwa', 'Odmianki', 'POWIAT', 'Powiaty', 'Parafie', 'SHG online']
        sheet_mejscowosci = create_sheet(wb=wb, title="Miejscowości", columns=point_sheet_columns)
        dim_holder = DimensionHolder(worksheet=sheet_mejscowosci)

        # for col in range(sheet_mejscowosci.min_column, sheet_mejscowosci.max_column + 1):
        col = 1 # shg_id
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet_mejscowosci, min=col, max=col, width=10)
        col = 2 # nazwa
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet_mejscowosci, min=col, max=col, width=20)
        col = 3 # odmianki
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet_mejscowosci, min=col, max=col, width=50)
        col = 4 # powiat
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet_mejscowosci, min=col, max=col, width=15)
        col = 5 # powiaty
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet_mejscowosci, min=col, max=col, width=25)
        col = 6 # parafie
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet_mejscowosci, min=col, max=col, width=20)
        col = 7 # shg online
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet_mejscowosci, min=col, max=col, width=40)

        sheet_mejscowosci.column_dimensions = dim_holder

        # słownik odnośników do ewentualnego uzupełnienia odmianek
        if os.path.exists(odnosniki_path):
            with open(odnosniki_path, 'r', encoding='utf-8') as o:
                lines = o.readlines()

            odnosniki = {}
            if lines:
                for line in lines:
                    line = line.strip()
                    tmp = line.split('|')
                    if tmp[0].upper() in odnosniki:
                        odnosniki[tmp[0].upper()].append(tmp[1])
                    else:
                        odnosniki[tmp[0].upper()] = [tmp[1]]

        # przetwarzanie wierszy
        with open(input_path, mode='r', encoding='utf-8') as csv_file:
            csv_reader = csv.DictReader(csv_file, delimiter='#')

            for row in csv_reader:
                p0 = p1= ''
                miejsce_id = row['id'].strip()
                nazwa = row["header"]

                # if not 'HUBNIN' in nazwa:
                #     continue

                # czyszczenie nazwy
                nazwa = re.sub(r'<.*?>', '', nazwa)
                if nazwa.endswith(','):
                    nazwa = nazwa[:-1]

                # punkt 0 - odmianki
                p0 = row['0'].strip()
                odmianki = ''
                p0 = remove_html(p0).strip()

                # dla słownika krakowskiego odcinamy treść p 0 po inf o własności
                if nr == 3:
                    pos = p0.find('Własn. ')
                    if pos != -1:
                        p0 = p0[:pos]
                elif nr == 7: # dla słownika poznańskiego nie ma sensu szukać po Uwaga:
                    pos = p0.find('Uwaga: ')
                    if pos != -1:
                        p0 = p0[:pos]

                # weryfikacja czy to miejscowość a nie odnośnik, polana, tenuta, kasztelania
                if not czy_miejscowosc(nazwa, p0):
                    # jeżeli nie to uzupełnianie słownika odnośników
                    if p0.startswith('→'):
                        tmp = p0.split('→')
                        for t in tmp:
                            if t.strip() != '':
                                with open(odnosniki_path, 'a', encoding='utf-8') as o:
                                    o.write(f'{t.strip()}|{nazwa}\n')
                    continue
                else:
                    # jeżeli wyglada na miejscowość ale są to wybrane słowniki np. Sanok
                    # to analiza czy jednak nie jest to odnośnik i wówczas uzupełnianie
                    # słownika odnośników
                    if nr == 10: # Wyszogród
                        if p0.startswith('zob. '):
                            p0 = p0.replace('zob. ', '')
                            if p0.endswith('.'):
                                p0 = p0[:-1]
                            with open(odnosniki_path, 'a', encoding='utf-8') as o:
                                o.write(f'{p0}|{nazwa}\n')

                            continue

                    elif nr in (8, 3): # Sanok, Kraków
                        if p0.startswith('→'):
                            tmp = p0.split('→')
                            for t in tmp:
                                if t.strip() != '':
                                    with open(odnosniki_path, 'a', encoding='utf-8') as o:
                                        o.write(f'{t.strip()}|{nazwa}\n')

                            continue

                # ----------------- START odmianki -----------------------------

                poznan_style = False
                patterns = [r'^\d{4}',
                            r'\[pocz\.\s+[XVI]+\s+w\.',
                            r'^ok\.\s+\d{4}',
                            r'\[\d{4}\d]\s+']
                for pattern in patterns:
                    match = re.search(pattern, p0)
                    if match:
                        poznan_style = True
                        break

                if poznan_style and nr in [7, 12, 13, 14]: # poznan, liw, czersk, podlasie
                    odmianki = get_odmianki_poznan()
                else:
                    odmianki = get_odmianki_standard()


                # --------------- END odmianki ---------------------------------

                # to powoduje błędy w przypadku powtarzających się nazw
                # if nazwa in odnosniki:
                #     for item_odn in odnosniki[nazwa]:
                #         if item_odn not in odmianki:
                #             odmianki += ', ' + item_odn
                #             #print(nazwa, ': ', item_odn)
                #        # else:
                #             #print('jest: ', item_odn)
                #     if odmianki.startswith(', '):
                #         odmianki = odmianki[2:]

                # powiaty i parafie
                p0 = p0.replace(r'\n', ' ')
                p1 = row['1'].strip()
                p1 = p1.replace(r'\n', ' ')

                pomin_powiaty_parafie = False
                if nr == 7 and 'przekazy o nieokreślonej lokalizacji' in p0:
                    pomin_powiaty_parafie = True

                p_powiat = p1
                if nr in [3]: # Kraków
                    krakow_odwrotnie = ['JANOWICE DOLNE i GÓRNE', 'KASINKA']
                    if nazwa not in  krakow_odwrotnie:
                        pos = p_powiat.find(' par.')
                        if pos != -1:
                            p_powiat = p_powiat[:pos]

                powiat = powiat_last = ''

                if not pomin_powiaty_parafie:
                    pattern = r'(pow|distr|Pow|Distr|districtus|dystr)\.?\s+[\w]+\.?'
                    match_pow = [x.group() for x in re.finditer(pattern, p_powiat)]
                    if match_pow:
                        match_pow = [expand_abbr(pow.strip(), ['pow.','Pow.','distr.','Distr.', 'districtus']) for pow in match_pow]
                        match_pow = [x for x in match_pow if len(x) > 1]
                        if match_pow:
                            powiat_last = match_pow[-1]
                            match_pow = list(set(match_pow))
                            powiat = ', '.join(match_pow)
                    else:
                        p_powiat = p0
                        if nr in [3]: # Kraków
                            pos = p_powiat.find(' par.')
                            if pos != -1:
                                p_powiat = p_powiat[:pos]

                        match_pow = [x.group() for x in re.finditer(pattern, p_powiat)]
                        if match_pow:
                            match_pow = [expand_abbr(pow.strip(), ['pow.','Pow.','distr.','Distr.']) for pow in match_pow]
                            match_pow = [x for x in match_pow if len(x) > 1]
                            if match_pow:
                                powiat_last = match_pow[-1]
                                match_pow = list(set(match_pow))
                                powiat = ', '.join(match_pow)

                # parafie
                parafia = ''
                if not pomin_powiaty_parafie:
                    pattern = r'[pP]{1}ar[\.,;]?\s+[→]?\s?[\w\.]+(\s?\w{0,})+'
                    match_par = [x.group() for x in re.finditer(pattern, p1)]
                    if match_par:
                        match_par = [clear_value(par.strip(), ['par.','par,', 'par;', 'par']) for par in match_par]
                        match_par = [parafia_nazwa(par, nazwa) for par in match_par]
                        match_par = [x for x in match_par if x]
                        match_par = unique(match_par)
                        parafia = ', '.join(match_par)
                        if '→' in parafia:
                            parafia = parafia.replace('→',' ')
                            parafia = double_space(parafia)
                    else:
                        match_par = [x.group() for x in re.finditer(pattern, p0)]
                        if match_par:
                            match_par = [clear_value(par.strip(), ['par.','par,', 'par;', 'par']) for par in match_par]
                            match_par = [parafia_nazwa(par, nazwa) for par in match_par]
                            match_par = [x for x in match_par if x]
                            match_par = unique(match_par)
                            parafia = ', '.join(match_par)
                            if '→' in parafia:
                                parafia = parafia.replace('→',' ')
                                parafia = double_space(parafia)
                        else:
                            pattern = r'[pP]{1}arafia\s+[łac\.]?\s?[\w\.]+(\s?\w{0,})+'
                            match_par = [x.group() for x in re.finditer(pattern, p1)]
                            if match_par:
                                match_par = [clear_value(par.strip(), ['parafia']) for par in match_par]
                                match_par = [parafia_nazwa(par, nazwa) for par in match_par]
                                match_par = [x for x in match_par if x]
                                match_par = unique(match_par)
                                parafia = ', '.join(match_par)
                                if '→' in parafia:
                                    parafia = parafia.replace('→',' ')
                                    parafia = double_space(parafia)

                if not pomin_powiaty_parafie:
                    if not parafia:
                        # wyjątki
                        wyjatki = ['par. Ś. Świerada w Tropiu', 'par. Ś. Szczepana w Krakowie']
                        for wyj in wyjatki:
                            if wyj in p1 or wyj in p0:
                                parafia = wyj
                                break
                    else:
                        if parafia.endswith('. Własn'):
                            parafia = parafia[:-7]

                # print('Powiat:', powiat)
                # print('Parafia:', parafia)

                akapit = tuple([miejsce_id, nazwa, odmianki, powiat_last, powiat, parafia, ''])
                sheet_mejscowosci.append(akapit)

        # usunięcie pustego pierwszego arkusza
        sheet1 = wb['Sheet']
        wb.remove(sheet1)

        # hyperlinki do SHG

        max_rows = sheet_mejscowosci.max_row
        for row in sheet_mejscowosci.iter_rows(2, max_rows):
            m_id = row[0].value
            row[6].hyperlink = f'http://www.slownik.ihpan.edu.pl/search.php?id={m_id}'
            row[6].value = f'http://www.slownik.ihpan.edu.pl/search.php?id={m_id}'
            row[6].style = "Hyperlink"

        # zapis pliku xlsx
        wb.save(filename=output_path)
