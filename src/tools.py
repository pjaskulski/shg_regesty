import re
import bs4
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
#import spacy


#nlp = spacy.load('pl_core_news_lg')
#nlp.max_length = 16000000

def load_odmiany(path) -> dict:
    """ funkcja ładuje słownik końcówek odmian wyrazów
    """
    odmiany = {}
    with open(path, 'r', encoding='utf-8') as f:
        lines = f.read().splitlines() 
        for line in lines:
            key,value = line.split(',')
            odmiany[key] = value

    return odmiany


def load_short(path) -> dict:
    """ funkcja ładuje słownik skrótów
    """
    slownik = {}
    wb_sl = openpyxl.load_workbook(path)
    ws_sl = wb_sl.active

    # nazwy kolumn w pierwszym arkuszu
    col_sl = {}
    nr_col = 0
    for column in ws_sl.iter_cols(1, ws_sl.max_column):
        col_sl[column[0].value] = nr_col
        nr_col += 1

    max = ws_sl.max_row

    for row in ws_sl.iter_rows(2, max):
        key = row[col_sl['skrot']].value
        value = row[col_sl['haslo']].value
        slownik[key] = value

    return slownik


def set_alignment(sheet):
    """ zmiana wyrównania w kolumnie Punkt 
    """
    for row in sheet[2:sheet.max_row]:
        cell = row[2]             
        cell.alignment = Alignment(horizontal='center')


def create_sheet(wb, title, columns):
    """ tworzy nowy arkusz zgodnie z parametrami
        wb - workbook
        title - tytuł arkusza
        columns - lista nazw kolumn
    """
    regest_sheet = wb.create_sheet(title=title)
    # nagłówki i style arkusza
    regest_sheet.append(columns)
    for col in range(1, len(columns) + 1):
        regest_sheet.cell(row = 1, column = col).font = Font(name='Calibri', size=11, bold = True)
    regest_sheet.freeze_panes = "A2"
    regest_sheet.print_title_rows = '1:1'
    
    if title == "Własność":
        regest_sheet.column_dimensions[get_column_letter(1)].width = 8      # Id
        regest_sheet.column_dimensions[get_column_letter(2)].width = 25     # Nazwa
        regest_sheet.column_dimensions[get_column_letter(3)].width = 45     # Własność   
    else:
        regest_sheet.column_dimensions[get_column_letter(1)].width = 8      # Id
        regest_sheet.column_dimensions[get_column_letter(2)].width = 25     # Nazwa
        regest_sheet.column_dimensions[get_column_letter(3)].width = 8      # Punkt
        #cell.alignment = Alignment(horizontal='center')
        regest_sheet.column_dimensions[get_column_letter(4)].width = 20     # Rok
        regest_sheet.column_dimensions[get_column_letter(5)].width = 40     # Treść
        regest_sheet.column_dimensions[get_column_letter(6)].width = 40     # Treść TAGI
        regest_sheet.column_dimensions[get_column_letter(7)].width = 20     # Źródła
        regest_sheet.column_dimensions[get_column_letter(8)].width = 20     # Osoby
        regest_sheet.column_dimensions[get_column_letter(9)].width = 20     # Miejsca
        regest_sheet.column_dimensions[get_column_letter(10)].width = 20    # Obiekty
        regest_sheet.column_dimensions[get_column_letter(11)].width = 40    # Regest
        regest_sheet.column_dimensions[get_column_letter(12)].width = 60    # Link
      
    return regest_sheet


def add_tags_to_content(doc, obiekty) -> str:
    """ funkcja taguje przekazany document spacy i zwraca otagowany tekst 
    """
    people_start = {}
    people_end = {}
    places_start = {}
    places_end = {}
    counter = 0
    text = ""

    # wyszukiwanie nazw własnych
    for ent in doc.ents:
        if ent.label_ == 'persName':
            people_start[ent.start] = '<persName>'
            people_end[ent.end] = '</persName>'
        elif ent.label_ == 'placeName' or ent.label_ == 'geoName':
            places_start[ent.start] = '<geoName>'
            places_end[ent.end] = '</geoName>'

    # tworzenie tekstu z tagami
    for token in doc:
        if counter in people_start.keys():
            text += people_start[counter]
            text += token.text_with_ws
        elif counter in places_start.keys():
            text += places_start[counter]
            text += token.text_with_ws
        elif counter in people_end.keys():
            jump = ''
            if text.endswith(' '):
                text = text[:-1]
                jump = ' '
            text += people_end[counter]
            if token.pos_ == 'NOUN' and token.lemma_ in obiekty:
                whitespace = ''
                if token.text_with_ws.endswith(' '):
                    whitespace = ' '
                text += jump + '<placeName>' + token.text + '</placeName>' + whitespace
            else:
                text += jump + token.text_with_ws 
        elif counter in places_end.keys():
            jump = ''
            if text.endswith(' '):
                text = text[:-1]
                jump = ' '
            text += places_end[counter]    
            if token.pos_ == 'NOUN' and token.lemma_ in obiekty:
                whitespace = ''
                if token.text_with_ws.endswith(' '):
                    whitespace = ' '
                text += jump + '<placeName>' + token.text + '</placeName>' + whitespace
            else:
                text += jump + token.text_with_ws
        else: 
            if token.pos_ == 'NOUN' and token.lemma_ in obiekty:
                whitespace = ''
                if token.text_with_ws.endswith(' '):
                    whitespace = ' '
                text += '<placeName>' + token.text + '</placeName>' + whitespace
            else:
                text += token.text_with_ws

        counter += 1

    if counter in people_end.keys():
        text += people_end[counter]
    elif counter in places_end.keys():
        text += places_end[counter]

    return text


def process_point(sheet, id, nazwa, point_num, point_text, obiekty, slownik, odmiany,
                    get_owner=0, 
                    get_year=True, 
                    get_content=True,
                    get_content_tag=True,
                    get_source=True,
                    get_persons=True,
                    get_places=True,
                    get_objects=True,
                    get_regest=True):
    """przetwarzanie treści punktu z zapisem do wskazanego arkusza
       zwraca własność lub pusty string
       sheet - wskazanie na arkusz
       id - identyfikator miejscowości
       nazwa - nazwa miejscowości
       point_num - numer punktu
       point_text - treść punktu
       obiekty - lista obiektów do wyszukania typu karczma, las, most
       get_owner - czy szukać własności (w zasadzie tylko w p. 3)
       get_year - czy szukać roku, lat
       get_content - czy tworzyć pole z treścią
       get_content_tag - czy tworzyć pole z treścią otagowaną?
       get_source - czy tworzyć pole ze źródłem
       get_persons - czy tworzyć pole z osobami
       get_places - czy tworzyć pole z miejscami
       get_objects - czy tworzyć pole z obiektami
       get_regest - czy tworzyć pole z regestem
    """
    ownership = old_year = ""
    reg_list = []
    point = point_num

    count_pers = 0
    count_geo = 0

    reg_list, ownership = get_regesty(point_text, get_owner)
    for item in reg_list:
        item = item.strip()
        if get_year:
            year = get_year_from_regest(item)
        if get_source:
            source = get_source_from_regest(item)
        if get_content and get_year and get_source:
            contents = get_contents_from_regest(item, year, source, nazwa)

        people = places = objects = ""
        contents_tag = ""

        if get_content and contents != "":
            contents = expand_shortcuts(contents, slownik, odmiany, nazwa)

        if get_content and (get_persons or get_places or get_objects) and contents != "":
            doc = nlp(contents)
            # wyszukiwanie nazw własnych
            for ent in doc.ents:
                if get_persons and ent.label_ == 'persName':
                    people += ent.text + ', '
                    count_pers += 1
                elif get_places and (ent.label_ == 'placeName' or ent.label_ == 'geoName'):
                    places += ent.text + ', '
                    count_geo += 1
            # wyszukiwanie obiektów typu karczma, wieś, droga, most
            if get_objects:
                for token in doc:
                    if token.pos_ == 'NOUN' and token.lemma_ in obiekty:
                        objects += token.lemma_ + ', '

            if people.endswith(', '):
                people = people[:-2]
            if places.endswith(', '):
                places = places[:-2]
            if objects.endswith(', '):
                objects = objects[:-2]

            # czyszczenie z duplikatów
            if get_objects:
                tmp_obj_list = objects.split(', ')
                tmp_set = set(tmp_obj_list)
                objects = ", ".join(tmp_set)

            if get_content_tag:
                contents_tag = add_tags_to_content(doc, obiekty)

        item = item.replace('[@', '[').replace('@]',']')
        if get_year:
            if year == "" and old_year != "":
                year = old_year

        tmp_regest = [id, nazwa, point]
        if get_year:
            tmp_regest.append(year)
        if get_content:
            tmp_regest.append(contents)
        if get_content_tag:
            tmp_regest.append(contents_tag)
        if get_source:
            tmp_regest.append(source)
        if get_persons:
            tmp_regest.append(people)
        if get_places:
            tmp_regest.append(places)
        if get_objects:
            tmp_regest.append(objects)
        if get_regest:
            tmp_regest.append(item)

        regest = tuple(tmp_regest)
        sheet.append(regest)
        if get_year:
            old_year = year

    ret_owner = find_owner(ownership)

    return ret_owner, count_pers, count_geo


# from: https://stackoverflow.com/questions/26808913/split-string-at-commas-except-when-in-bracket-environment/26809037
def semisplit(s):
    """ funkcja dzieli przekazany tekst na wiersze według średnika pomijając
        średniki w nawiasach, zwraca listę
    """
    parts = []
    bracket_level = 0
    current = []
    # trick to remove special-case of trailing chars
    for c in (s + ";"):
        if c == ";" and bracket_level == 0:
            parts.append("".join(current))
            current = []
        else:
            if c == "(":
                bracket_level += 1
            elif c == ")":
                bracket_level -= 1
            current.append(c)
    return parts


def clear_regest(item) -> str:
    """ czyszczenie regestu ze znaków specjalnych, tagów html itp.
    """
    # dodatkowe miejsce podziału ze względu na błędy w słowniku (np. podpunkt 3A sklejony
    # z poprzednim)
    item = re.sub(r'\.<\/p>(\\n|\\\\n)<p>', '; ', item)

    item = item.replace(r'\n','').replace('</p>','').replace('<b>','').replace('</b>','').strip()
    item = item.replace('<i>','').replace('</i>','').strip()
    # w regeście może być więcej niż 1 nawias a trzeba sprawdzić czy nie brakuje
    # nawiasu zamykającego
    if item.count('(') != item.count(')'):
        item += ")"
    if item.endswith(")."):
        item = item[:-1]
    item = re.sub('<span[\w\s="]+>\s?<\/span>', '', item)
    return item


def get_ownership(tmp) -> tuple:
    """ funkcja zwraca typ własności i resztę zapisu do regestu nr 1
    """

    ownership = tmp_regest_1 = ""
    patterns = [r'\.\s+\[?\d{4}', r'\d{4}', r'\.\s+\[']

    for pattern in patterns: 
        match = re.search(pattern, tmp)
        if match:
            ownership = tmp[:match.start()]
            tmp_regest_1 = tmp[match.start()+1:]
            break

    if ownership == "":
        stop_words = ['Treść punktu', 'Sprawy własnościowe', '. Przed', ', od']
        for word in stop_words:
            pos = tmp.find(word)
            if pos:
                ownership = tmp[:pos]
                tmp_regest_1 = tmp[pos:]
                break

    if ownership == "":
        pattern = r'\d{4}\s+własn\.\s+szlach\.'
        match = re.search(pattern, tmp)
        if match:
            pos = match.end()
            pattern = r',\s+\d{4}'
            match = re.search(pattern, tmp)
            if match:
                ownership = tmp[:match.start()]
                tmp_regest_1 = tmp[pos+1:]

    return ownership, tmp_regest_1


def get_regesty(punkt, find_ownership = 0) -> list:
    """ funkcja wydobywa listę regestów z przekazanego tekstu punktu,
        opcjonalnie wydobywa też rodzja własności np dla punktu 3
        punkt - treść punktu
        find_ownership - czy wydobywać własność, domyślnie nie
    """

    # jeżeli przypis (może być wiele wiele)
    footnote_text = {}
    if 'footnoteAnchor' in punkt:
        footnoteSoup = bs4.BeautifulSoup(punkt, "html.parser")
        anchors = footnoteSoup.find_all("a", class_="footnoteAnchor")
        if len(anchors) > 0:
            for anchor in anchors:
                an = anchor.text.strip()
                span = footnoteSoup.find("span", id="footnote"+an)
                if span != None:
                    footnote_text[an] = span.text.strip()
                    anchor.string = f'({an})'
                    anchor.unwrap()
                    span.decompose() 
        punkt = footnoteSoup.get_text()

    # wstępny podział
    lista = punkt.split(');')            

    # dodatkowe czyszczenie
    p_lista = []
    for item in lista:
        # czyszczenie regestu ze spacji, znaczników html itp. uzupełnianie brakujących
        # nawiasów
        item = clear_regest(item)

        # jeżeli były przypisy to zamiana tymczasowych znaczników na treść przypisu
        if len(footnote_text) > 0:
            for key in footnote_text.keys():
                tmp = f'({key})'
                if tmp in item:
                    item = item.replace(tmp, f'[@przyp. {key}:{footnote_text[key]}@]')

        if ";" in item:
            # dalsze dzielenie, po wzorcu ale też ze sprawdzeniem czy nie w nawiasie 
            s_list = []
            pattern = ';\s+\d{4}'
            subitems = re.finditer(pattern, item)
            for subitem in subitems: 
                is_bracket = False
                for i in range(subitem.start(),0,-1):
                    if item[i] == "(":
                        is_bracket = True
                        break
                if not is_bracket:
                    s_list.append(subitem.start())

            # był wzorzec typu '; 1425' (lub wiele)
            if len(s_list) > 0:
                pos = 0
                for s in s_list:
                    p_lista.append(item[pos:s].strip())
                    pos = s + 1
                p_lista.append(item[pos:])
            else:   
                p_lista.append(item)
        else:
            p_lista.append(item)


    # pierwszy zapis np. w p. 3 być może trzeba podzielić i wydobyć własność:
    ownership = ""
    if find_ownership == 1:
        ownership, regest_1 = get_ownership(p_lista[0])
        if ownership != "":
            p_lista[0] = regest_1

    return p_lista, ownership


def get_text_with_year(text: str) -> str:
    """ zwraca początek tekstu z rokiem, latami"""
    result = ''
    znaki = '0123456789- ,aAn.[]'
    for i in text:
        if i in znaki:
            result += i
        else:
            break

    return result


def get_year_from_regest(text) -> str:
    """zwraca rok lub lata z regestu
    """
    # if '1442n. pow. sand. (Mp. IV 1434)' in text:
    #     print()

    result = ""
    patterns = [
                r'\[\d{4}\]',
                r'\[\d{4}\s?-\s?\d{1,4}(,\s+\d{4}\s?-\s?\d{1,4})+\]',
                r'\[\d{4}\s?-\s?\d{1,2}\]',

                r'ok\.\s+\d{4}',
                r'^a\.\s+\d{4}\s?-\s?\d{1}(n\.){0,1}',
                r'^a\.\s+\d{4}(n\.){0,1}',
                r'^A\.\s+\d{4}(,\s+\d{4}(n\.){0,1})*',
                r'^\d{4}(n\.){0,1}'
                r'\d{4}-przed\s+\d{4}',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){2})*',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){3})*',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){4})*',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){5})*',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){6})*',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){7})*',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){8})*',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){9})*',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){10})*',
                r'\d{4}((,\s+\d{4}(n\.){0,1}){11})*',
                r'\d{4}\s?-\s?\d{2},\s+\d{4},\s+\d{4},\s+\d{4},\s+\d{4}\s?-\s?\d{1},\s+\d{4},\s+\d{4}\s?-\s?\d{1}(n\.){0,1}',
                r'\d{4},\s+\d{4},\s+\d{4}-\d{1},\s+\d{4}-\d{1}(,\s+\d{4}(n\.){0,1})*',
                r'\d{4}(,\s+[\d-]{4,7}(n\.){0,1})*',
                r'\d{4},\s+\d{4}-\d{1},\s+\d{4},\s+\d{4}-\d{1}(n\.){0,1}',
                r'\d{4}-\d{2},\s+\d{4}-\d{1}(,\s+\d{4}(n\.){0,1})*',
                r'\(\d{4}\s?-\s?\d{1}\)\s+\d{4}(n\.){0,1}',
                r'\d{4}\s?-\s?\d{2},\s+\d{4},\s+\d{4}(n\.){0,1}',
                r'\d{4}(n\.){1},\s+\d{4}-\d{2}(n\.){1}',
                r'\d{4}\/\d{1},\s+\d{4},\s+\d{4}',
                r'\d{4},\s+\d{4},\s+\d{4}',
                r'\d{4}-\d{2},\s+\d{4}-\d{1},\s+\d{4}-\d{1},\s+\d{4}-\d{1},\s+\d{4}-\d{1}',
                r'\d{4}-\d{1},\s+\d{4}-\d{1},\s+\d{4}-\d{1}',
                r'\d{4},\s+\d{4}\s?-\s?\d{2},\s+\d{4},\s+\d{4}(n\.){0,1}',
                r'\d{4},\s+\d{4}\s?-\s?\d{2},\s+\d{4},\s+\d{4},\s+\d{4}(n\.){0,1}',
                r'\d{4},\s+\d{4}\s?-\s?\d{2}(n\.){0,1}',
                r'\d{4},\s+\d{4},\s+\d{4}-\d{2}(n\.){0,1}',
                r'\d{4},\s+\d{4}\s?-\s?\d{1}(n\.){0,1}',
                r'\d{4}\s?-\s?\d{1},\s+\d{4}\s?-\s?\d{1}(n\.){0,1}',
                r'\d{4},\s+\d{4},\s+\d{4}-\d{1}(,\s+\d{4}(n\.){0,1})*',
                r'\d{4}-\d{2},\s+\d{4}-\d{4},\s+\d{4}(n\.){0,1}',
                r'\d{4}\s?-\s?\d{2},\s+\d{4}(n\.){0,1}',
                r'\d{4}\s?-\s?\d{1},\s+\d{4}(n\.){0,1}',
                r'\d{4}\s?-\s?\d{2},\s+\d{4}\s?-\s?\d{4}(n\.){0,1}',
                r'\d{4}\s?-\s?\d{2},\s+\d{4}\s?-\s?\d{1}(n\.){0,1}',
                r'\d{4},\s*\d{4}(n\.){0,1}',
                r'\d{4}\s?-\s?\d{4}(n\.){0,1}',
                r'\d{4}\s?-\s?\d{2}(n\.){0,1}',
                r'\d{4}\s?-\s?\d{1}(n\.){0,1}',
                r'\d{4}\d{4}(n\.){0,1}',
                r'\d{4}-\d{3}(n\.){0,1}'
               ]

    for pattern in patterns:
        # pierwszych 25 znaków
        #first_25 = text[:25]
        begin_text = get_text_with_year(text)
        #pos = first_25.find('zm.')
        #if pos != -1:
        #    first_25 = first_25[:pos]

        match = re.search(pattern, begin_text)
        temp = ""
        if match:
            temp = match.group()
        if len(temp) > len(result):
            result = temp.strip()

    return result


def get_source_from_regest(text) -> str:
    """zwraca sygnatury źródeł
    """
    result = ""
    # przy szukaniu sygnatur źródeł pomijanie przypisów
    if '[@' in text:
        text = text[:text.rfind("[@")] + text[text.rfind('@]')+2:]

    # wycinanie zawartości ostatniego nawiasu
    if "(" in text:
        result = text[text.rfind("(")+1:text.rfind(")")]

    # lub wszystkie nawiasy
    #pattern = r'\(.*?\)'
    #matches = [x.group().replace('(', '').replace(')', '') for x in re.finditer(pattern, text)]
    #if matches:
    #    result = ', '.join(matches)

    return result


def get_contents_from_regest(item, year, source, nazwa) -> str:
    """ zwraca samą treść regestu, bez roku i źródeł
    """

    if source != "":
        #item = item.replace('(' + source + ')', '')
        item = re.sub(r'\(.*?\)', '', item)
    if year != "":
        item = item.replace(year, '')
    #if '[@' in item:
    #    item = item[:item.find('[@')]

    return item.replace('[]','').strip()


def expand_shortcuts(text, slownik, odmiany, nazwa) -> str:
    """ funkcja rozwija znane skróty w przekazanym tekście 
    """
    first = nazwa[0].upper() + "."
    nazwa = nazwa[0].upper() + nazwa[1:]

    s_odmiany = sorted(list(odmiany.keys()), key=len, reverse=True)
    if ' z ' + first in text:
        find = False 
        for key in s_odmiany:
            if nazwa.endswith(key):
                nazwa = nazwa[:-1*len(key)] + odmiany[key]
                find = True
                break
        if find:
            text = text.replace(' z ' + first, ' z ' + nazwa)
        else:
            text = text.replace(first, nazwa)
    else:
        text = text.replace(first, nazwa)

    # skróty słownikowe na razie nie rozwijane
    # tmp_keys = sorted(list(slownik.keys()), key=len, reverse=True)
    # for key in tmp_keys:
    #     text = text.replace(' '+key+' ', ' '+slownik[key]+' ')
    #     text = text.replace(' '+key+']', ' '+slownik[key]+']')
    #     text = text.replace(' '+key+';', ' '+slownik[key]+';')
    #     text = text.replace(' '+key+',', ' '+slownik[key]+',')

    #     if text.endswith(' '+key):
    #         text = text.replace(' '+key, ' '+slownik[key])
    #     elif text.startswith(key+' '):
    #         text = text.replace(key+' ', slownik[key]+' ')

    return text


def create_hyperlink(sheet):
    """ funkcja tworzy linki do strony słownika historycznego 
         ale tylko dla pierwszego regestu z danej miejscowości
    """
    sheet_max = sheet.max_row
    old = ""
    for i in range(2, sheet_max):
        if sheet[f'A{i}'].value != old:
            sheet[f'D{i}'].hyperlink = sheet[f'D{i}'].value
            sheet[f'D{i}'].style = "Hyperlink"
            old = sheet[f'A{i}'].value
        else:
            sheet[f'D{i}'].value = ""


def find_owner(ownership) -> str:
    """zwraca string z listą rodzajów własności w formie tekstu 
       rodzielonego przecinkami
    """
    ret_owner = ""
    if ownership != "":
        if 'szlach' in ownership:
            ret_owner += 'szlachecka,'
        if 'kl.' in ownership:
            ret_owner += 'klasztorna,'
        if 'król.' in ownership or 'król' in ownership or 'monarsza' in ownership or 'królowych' in ownership:
            ret_owner += 'królewska,'
        if 'bpa' in ownership or 'arcbpa' in ownership or 'bpów' in ownership or 'arcbpstwa' in ownership or 'bpstwa' in ownership:
            ret_owner += 'biskupia,'
        if 'książęca' in ownership or 'ks.' in ownership or 'książąt' in ownership or 'księcia' in ownership or 'księżnych' in ownership or 'księżnej' in ownership:
            ret_owner += 'książęca,'
        if 'kościoła' in ownership or 'kolegiat' in ownership or 'kol.' in ownership or 'dziekanii' in ownership or 'prebendy' in ownership or 'preb' in ownership or 'duch.' in ownership:
            ret_owner += 'kościelna,'
        if 'Własn. m.' in ownership or 'miejska' in ownership or 'miasta' in ownership or 'mieszczan' in ownership or 'rajców' in ownership:
            ret_owner += 'miejska,'
        if 'kap. krak.' in ownership or 'kapit. krak.' in ownership:
            ret_owner += 'kapitulna,'
        if 'kasztelanii' in ownership:
            ret_owner += 'kasztelanii,'
        if 'rycerska' in ownership or 'ryc.' in ownership:
            ret_owner += 'rycerska,'
        if 'wójtostwa' in ownership:
            ret_owner += 'wójtostwa,'

    # wielkorządców?
    # stwa ?
        if ret_owner.endswith(','):
            ret_owner = ret_owner[:-1]

    return ret_owner
