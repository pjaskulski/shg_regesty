""" Analiza występowania i tagowania jednostek monetarnych - dane z xlsx """

import re
import os
from pathlib import Path
import spacy
import openpyxl
from spacy.matcher import Matcher
from monety import false_coins_list, rule_patterns


slowniki = {1: 'Benedyktyni',
            2: 'Chełmno',
            3: 'Kraków',
            4: 'Lublin',
            5: 'Lublin_zaginione',
            6: 'Płock',
            7: 'Poznań',
            8: 'Sanok',
            9: 'Wieluń',
            10: 'Wyszogród',
            11: 'Warszawa',
            12: 'Liw'
}


def clear_text(text: str) -> str:
    """ czyści tekst z tagów html, odnośników bibliograficznych"""
    # nawiasy
    pattern = r'[\(\[].*?[\)\]]'
    text = re.sub(pattern, '', text)
    pattern = r'<.*?>'
    text = re.sub(pattern, '', text)
    text = text.replace(r'\n', ' ')

    return text


if __name__ == "__main__":
    patterns = rule_patterns()
    false_coins = false_coins_list()

    nlp = spacy.load("pl_core_news_lg")
    matcher = Matcher(nlp.vocab)
    matcher.add("Monety", patterns=patterns)

    for nr in range(1, 13):
        print(f"Słownik: {slowniki[nr]} ({nr})")
        lista = []

        input_path = Path('.').parent / f'output/hasla_{nr}_pkt.xlsx'
        output_path = Path('.').parent / f'output/monety_{slowniki[nr]}.txt'
        text_path = Path('.').parent / f'output/text_{slowniki[nr]}.txt'

        if os.path.exists(text_path):
            os.remove(text_path)

        wb_hasla = openpyxl.load_workbook(input_path)
        ws_hasla = wb_hasla.active

        # kolumny w źródłowym pierwszym arkuszu
        col_names = {}
        nr_col = 0
        for column in ws_hasla.iter_cols(1, ws_hasla.max_column):
            col_names[column[0].value] = nr_col
            nr_col += 1

        max_hasla = ws_hasla.max_row

        licznik = 0
        for row in ws_hasla.iter_rows(2, max_hasla):
            licznik += 1
            print(f'Przetwarzam wiersz {licznik} z {max_hasla}.')
            # p_full = []
            # p0 = row[col_names[0]].value
            # p1 = row[col_names[1]].value
            # p2 = row[col_names[2]].value
            # p3 = row[col_names[3]].value
            # p4 = row[col_names[4]].value
            # p5 = row[col_names[5]].value
            # p6 = row[col_names[6]].value
            # p8 = row[col_names[8]].value
            # p9 = row[col_names[9]].value
            
            p_full = ''
            body = row[col_names['body']].value
            if body:
                p_full = clear_text(body)

            # do analizy treść p. 1-6, oraz nagłówek i uwagi
            # if p0:
            #     p_full.append(clear_text(p0))
            # if p1:
            #     p_full.append(clear_text(p1))
            # if p2:
            #     p_full.append(clear_text(p2))
            # if p3:
            #     p_full.append(clear_text(p3))
            # if p4:
            #     p_full.append(clear_text(p4))
            # if p5:
            #     p_full.append(clear_text(p5))
            # if p6:
            #     p_full.append(clear_text(p6))
            # if p9:
            #     p_full.append(clear_text(p9))

            # z zapisem analizowanej treści do dodatkowego pliku
            with open(text_path, "a", encoding='utf-8') as f_text:
                #for point in p_full:
                if p_full:
                    f_text.write(f'{p_full}\n\n')
                    doc = nlp(p_full)
                    matches = matcher(doc)

                    # tylko nie zawierające się znaleziska
                    spans = [doc[start:end] for _, start, end in matches]
                    for span in spacy.util.filter_spans(spans):
                        moneta = span.text
                        moneta = re.sub(r'\d', '', moneta)
                        moneta = moneta.replace('/','').replace('½', '').replace(';','').strip()
                        moneta = moneta.replace('–','').replace(':', '')
                        if moneta not in false_coins:
                            lista.append(moneta)

        # zapis pliku z wynikami dla danego słownika
        with open(output_path, "w", encoding='utf-8') as f:
            lista = list(set(lista))
            for item in sorted(lista, key=str.casefold):
                f.write(f'{item}\n')
