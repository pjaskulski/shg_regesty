""" harmonizacja Słownika Historyczno-Geograficznego z AHP """

import sqlite3
from sqlite3 import Error
from pathlib import Path
import openpyxl
#from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
#from openpyxl.utils import get_column_letter

harmonizacja = {}
harm_shg_id = harm_shg_id_zgodne = 0

slowniki = {1: 'Benedyktyni',
            2: 'Chełmno',
            3: 'Kraków',
            4: 'Lublin',
            5: 'Lublin_zaginione',
            6: 'Płock',
            7: 'Poznań',
            8: 'Sanok',
            9: 'Wieluń',
            10: 'Wyszogród',
            11: 'Warszawa',
            12: 'Liw',
            13: 'Czersk'
}


def create_connection(db_file):
    """ create a database connection to the SQLite database
        specified by the db_file
    :param db_file: database file
    :return: Connection object or None
    """

    conn = None
    try:
        conn = sqlite3.connect(db_file)
        conn.enable_load_extension(True)
        conn.load_extension("./fuzzy.so")
        conn.load_extension("./spellfix.so")
        conn.load_extension("./unicode.so")


    except Error as sql_error:
        print(sql_error)

    return conn


def field_strip(value) -> str:
    """ funkcja przetwarza wartość pola z bazy/arkusza """
    if value:
        value = value.strip()
    else:
        value = ''

    return value


def field_list_strip(value) -> str:
    """ funkcja przetwarza wartość pola z bazy/arkusza """
    if value:
        value = value.split(',')
        value = [x.strip() for x in value]
    else:
        value = []

    return value


def get_ahp_id(shg_id, shg_nazwa, shg_powiat, shg_powiaty, shg_parafie, shg_odmianki) -> str:
    """ wyszukuje w ahp zbiorcza i zwraca id lub pusty string """
    result = ''

    # nazwa
    sql = f""" SELECT id_miejscowosci, nazwa_16w, nazwa_odmianki,
                     parafia, powiat_p, wojewodztwo, nazwa_niemiecka, sgh_id
              FROM zbiorcza
              WHERE
              LOWER(nazwa_16w) = '{shg_nazwa.strip().lower()}'
          """
    result = select_from_db(sql, shg_nazwa, shg_powiat, shg_powiaty, shg_parafie, shg_id)

    # nazwa niemiecka
    if not result:
        sql = f""" SELECT id_miejscowosci, nazwa_16w, nazwa_odmianki,
                     parafia, powiat_p, wojewodztwo, nazwa_niemiecka, sgh_id
              FROM zbiorcza
              WHERE
              LOWER(nazwa_niemiecka) = '{shg_nazwa.strip().lower()}'
          """
        result = select_from_db(sql, shg_nazwa, shg_powiat, shg_powiaty, shg_parafie, shg_id)

    # nazwa współczesna?
    if not result:
        sql = f""" SELECT id_miejscowosci, nazwa_16w, nazwa_odmianki,
                    parafia, powiat_p, wojewodztwo, nazwa_niemiecka, sgh_id
                FROM zbiorcza
                WHERE
                LOWER(nazwa_wspolczesna) = '{shg_nazwa.strip().lower()}'
                or unaccent(LOWER(nazwa_wspolczesna)) = '{shg_nazwa.strip().lower()}'
            """
        result = select_from_db(sql, shg_nazwa, shg_powiat, shg_powiaty, shg_parafie, shg_id)

    # odmianki
    if not result and shg_odmianki:
        for item in shg_odmianki:
            sql = f""" SELECT id_miejscowosci, nazwa_16w, nazwa_odmianki,
                        parafia, powiat_p, wojewodztwo, nazwa_niemiecka, sgh_id
                    FROM zbiorcza
                    WHERE
                    LOWER(nazwa_16w) = '{item.strip().lower()}'
                    or
                    LOWER(nazwa_odmianki) = '{item.strip().lower()}'
                """
            result = select_from_db(sql, shg_nazwa, shg_powiat, shg_powiaty, shg_parafie, shg_id)
            if result:
                break

    # levenshtein
    if not result:
        sql = f""" SELECT id_miejscowosci, nazwa_16w, nazwa_odmianki,
                     parafia, powiat_p, wojewodztwo, nazwa_niemiecka, sgh_id
              FROM zbiorcza
              WHERE
              levenshtein(translit(LOWER(nazwa_16w)), translit('{shg_nazwa.strip().lower()}')) < 2
          """
    result = select_from_db(sql, shg_nazwa, shg_powiat, shg_powiaty, shg_parafie, shg_id)


    return result


def select_from_db(sql, shg_nazwa, shg_powiat, shg_powiaty, shg_parafie, shg_id):
    """ przetwarza zapytanie """
    global harm_shg_id, harm_shg_id_zgodne
    result = ''

    cur = db_ahp.cursor()
    cur.execute(sql)
    rows = cur.fetchall()

    pasuje = best_pasuje = 0
    best_ahp = ''
    best_ahp_shg_id = ''

    if rows and len(rows) > 1:
        for row in rows:
            ahp_id_m = field_strip(row[0])
            ahp_powiat = field_strip(row[4])
            ahp_parafia = field_strip(row[3])
            ahp_nazwa16w = field_strip(row[1])
            ahp_odmianki = field_strip(row[2])
            ahp_nazwa_niem = field_strip(row[6])
            ahp_shg_id = field_strip(row[7])

            pasuje = 1 # zgodna nazwa
            # powiaty
            if shg_powiat and shg_powiat.lower() == ahp_powiat.lower():
                pasuje += 1
            else:
                if shg_powiaty:
                    for item in shg_powiaty:
                        if item.lower() in ahp_powiat.lower():
                            pasuje += 1
                            break

            # parafie
            if shg_parafie and ahp_parafia:
                for item in shg_parafie:
                    # jeżeli parafia własna to zmiana na nazwę
                    if item.strip().lower() == 'własna':
                        item = shg_nazwa
                    if item.lower() in ahp_parafia.lower():
                        pasuje += 1
                        break

            if pasuje > best_pasuje:
                best_ahp = ahp_id_m
                best_ahp_shg_id = ahp_shg_id
                best_pasuje = pasuje

        if best_pasuje > 1:
            result = best_ahp
            if best_ahp_shg_id:
                harm_shg_id += 1
                if best_ahp_shg_id == shg_id:
                    harm_shg_id_zgodne += 1
                else:
                    print(shg_nazwa, shg_id, best_ahp_shg_id)
    # jeżeli tylko jedna miejscowość o tej nazwie to weryfikacja czy parafia lub powiat się zgadzają
    elif rows and len(rows) == 1:
        row = rows[0]
        ahp_id_m = field_strip(row[0])
        ahp_powiat = field_strip(row[4])
        ahp_parafia = field_strip(row[3])
        ahp_nazwa16w = field_strip(row[1])
        ahp_odmianki = field_strip(row[2])
        ahp_nazwa_niem = field_strip(row[6])
        ahp_shg_id = field_strip(row[7])
        match_powiat = False
        if shg_powiat.lower() == ahp_powiat.lower():
            match_powiat = True
        elif shg_powiaty:
            for item in shg_powiaty:
                if item.lower() in ahp_powiat.lower():
                    match_powiat = True
                    break
        match_parafia = False
        if shg_parafie and ahp_parafia:
            for item in shg_parafie:
                # jeżeli parafia własna to zmiana na nazwę
                if item == 'własna':
                    item = shg_nazwa
                if item.lower() in ahp_parafia.lower():
                    match_parafia = True
                    break
        if match_powiat or match_parafia:
            result = ahp_id_m
        else:
            if not shg_powiaty and not shg_parafie:
                # jeżeli brak informacji o powiatach i parafiach ale jest tylko jedna
                # miejscowość zgodna z ahp to skrypt akceptuje wynik
                result = ahp_id_m

        if result:
            if ahp_shg_id:
                harm_shg_id += 1
                if ahp_shg_id == shg_id:
                    harm_shg_id_zgodne += 1
                else:
                    print(shg_nazwa, shg_id, best_ahp_shg_id)

    return result


# -------------- Główny program ------------------------------------------------
if __name__ == '__main__':

    for nr in range(1, 14):
        print(f"Słownik: {slowniki[nr]} ({nr})")

        input_path = Path('.').parent / f'shg_ahp/gotowe/shg_ahp_{slowniki[nr]}.xlsx'
        output_path = Path('.').parent / f'shg_ahp/gotowe_ahp/shg_ahp_{slowniki[nr]}.xlsx'
        ahp_path = Path('.').parent / 'data/ahp_zbiorcza.sqlite'

        db_ahp = create_connection(ahp_path)

        wb = openpyxl.load_workbook(input_path)
        ws = wb.active

        # kolumny w źródłowym pierwszym arkuszu
        col_names = {}
        nr_col = 0
        for column in ws.iter_cols(1, ws.max_column):
            col_names[column[0].value] = nr_col
            nr_col += 1

        max_hasla = ws.max_row

        licznik = znalezione = 0

        for row in ws.iter_rows(2, max_hasla):
            licznik += 1
            #print(f'Przetwarzam wiersz {licznik} z {max_hasla}.')

            # identyfikator shg z bazy online
            shg_id = field_strip(row[col_names['SHG_ID']].value)     # str
            # nazwa miejscowości
            nazwa = field_strip(row[col_names['Nazwa']].value)
            # lista odmianek
            odmianki = field_list_strip(row[col_names['Odmianki']].value)
            # lista parafii
            parafie = field_list_strip(row[col_names['Parafie']].value)
            # najnowszy powiat
            powiat = field_strip(row[col_names['POWIAT']].value)
            # lista powiatów
            powiaty = field_list_strip(row[col_names['Powiaty']].value)

            ahp_id = get_ahp_id(shg_id, nazwa, powiat, powiaty, parafie, odmianki)
            if ahp_id != '':
                znalezione += 1

            row[col_names['AHP']].value = ahp_id

        harmonizacja[slowniki[nr]] = znalezione/licznik

        # zapis pliku xlsx
        wb.save(filename=output_path)

    # podsumowanie
    for key_slownik, value_slownik in harmonizacja.items():
        print(f'{key_slownik}: {value_slownik*100:.2f}')

    print(f'Zgodność shg_id: {harm_shg_id_zgodne}/{harm_shg_id}')
