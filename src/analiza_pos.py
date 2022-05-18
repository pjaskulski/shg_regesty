""" rzeczowniki, czasowniki """

import re
from collections import Counter
from pathlib import Path
import openpyxl
import spacy
from tools import create_sheet


nlp = spacy.load('pl_core_news_lg')
nlp.max_length = 16000000


slowniki = {1: 'Benedyktyni',
            2: 'Chełmno',
            3: 'Kraków',
            4: 'Lublin',
            5: 'Lublin_zaginione',
            6: 'Płock',
            7: 'Poznań',
            8: 'Sanok',
            9: 'Wieluń',
            10: 'Wyszogród',
            11: 'Warszawa',
            12: 'Liw'
}

def has_numbers(input_string: str):
    """ czy liczba w tekście? """
    return any(char.isdigit() for char in input_string)


def first_cap(input_string: str) -> bool:
    """ czy pierwsza duża litera? """
    return input_string[0].isupper()


def clear_text(text: str) -> str:
    """ czyści tekst z tagów html, odnośników bibliograficznych"""
    # nawiasy
    pattern = r'[\(\[].*?[\)\]]'
    text = re.sub(pattern, '', text)
    pattern = r'<.*?>'
    text = re.sub(pattern, '', text)
    text = text.replace(r'\n', ' ')
    # skróty
    text = ' '.join([skroty[w] if w in skroty else w for w in text.split()])
    # krótkie słowa < 3 znaków
    text = ' '.join([w for w in text.split() if len(w)>=3])
    # słowa z liczbami
    text = ' '.join([w for w in text.split() if not has_numbers(w)])

    return text


skroty = {}
skr_path = Path('.').parent / 'data/skroty_analiza.xlsx'
wb_skr = openpyxl.load_workbook(skr_path)
ws_skr = wb_skr.active
for row in ws_skr.iter_rows(2, ws_skr.max_row):
    skrot = row[0].value
    slowo = row[1].value
    skroty[skrot] = slowo

total_nouns = []
total_verbs = []
total_nouns_text = {}
total_verbs_text = {}

total_noun_counter = Counter()
total_verb_counter = Counter()

for nr in range(1, 13):
    print(f"Słownik: {slowniki[nr]} ({nr})")

    output_path = Path('.').parent / f'output/wyniki/slowa_{slowniki[nr]}.xlsx'
    input_path = Path('.').parent / f'output/wyniki/hasla_{nr}_pkt.xlsx'

    nouns = []
    verbs = []
    nouns_text = {}
    verbs_text = {}

    noun_counter = Counter()
    verb_counter = Counter()

    wb_hasla = openpyxl.load_workbook(input_path)
    ws_hasla = wb_hasla.active

    wb = openpyxl.Workbook()
    ws = wb.active

    noun_sheet = create_sheet(wb=wb, title="Rzeczowniki", columns=['Rzeczownik', 'Frekwencja', '', 'Słowa'])
    verb_sheet = create_sheet(wb=wb, title="Czasowniki", columns=['Czasowniki', 'Frekwencja', '', 'Słowa'])

    # kolumny w źródłowym pierwszym arkuszu
    col_names = {}
    nr_col = 0
    for column in ws_hasla.iter_cols(1, ws_hasla.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    max_hasla = ws_hasla.max_row

    licznik = 0
    for row in ws_hasla.iter_rows(2, max_hasla):
        licznik += 1
        print(f'Przetwarzam wiersz {licznik} z {max_hasla}.')
        p_full = ''
        p0 = row[col_names[0]].value    
        p1 = row[col_names[1]].value
        p2 = row[col_names[2]].value
        p3 = row[col_names[3]].value
        p4 = row[col_names[4]].value
        p5 = row[col_names[5]].value
        p6 = row[col_names[6]].value
        p8 = row[col_names[8]].value
        p9 = row[col_names[9]].value

        # słownik 'Lubelskie zaginione' zawiera treść tylko w nagłówku
        if nr == 5 and p0:
            p_full += ' ' + clear_text(p0)

        # w pozostałych słownikach treść jest w p. 1-8
        if p1:
            p_full += ' ' + clear_text(p1)
        if p2:
            p_full += ' ' + clear_text(p2)
        if p3:
            p_full += ' ' + clear_text(p3)
        if p4:
            p_full += ' ' + clear_text(p4)
        if p5:
            p_full += ' ' + clear_text(p5)
        if p6:
            p_full += ' ' + clear_text(p6)
        if p8:
            p_full += ' ' + clear_text(p8)
        if p9:
            p_full += ' ' + clear_text(p9)

        #print(p_full)
        cdoc = nlp(p_full.strip())

        for token in cdoc:
            if token.pos_ == "NOUN":
                lemma = token.lemma_
                word = token.text
                if not first_cap(lemma):
                    nouns.append(lemma)
                    if lemma in nouns_text:
                        nouns_text[lemma].append(word)
                    else:
                        nouns_text[lemma] = [word]
            elif token.pos_ == "VERB":
                lemma = token.lemma_
                word = token.text
                if not first_cap(lemma):
                    verbs.append(token.lemma_)
                    if lemma in verbs_text:
                        verbs_text[lemma].append(word)
                    else:
                        verbs_text[lemma] = [word]

    # uzupełnianie total
    total_nouns += nouns
    total_verbs += verbs
    total_nouns_text.update(nouns_text)
    total_verbs_text.update(verbs_text)

    # frekwencja rzeczowników i czasowników
    noun_counter.update(Counter(nouns))
    verb_counter.update(Counter(verbs))

    tmp_nouns = list(noun_counter.most_common())
    tmp_verbs = list(verb_counter.most_common())

    for tmp in tmp_nouns:
        if tmp[0] in nouns_text:
            lista = list(set(nouns_text[tmp[0]]))
            tmp = tmp + ('', ', '.join(lista),)
        noun_sheet.append(tmp)

    for tmp in tmp_verbs:
        if tmp[0] in verbs_text:
            lista = list(set(verbs_text[tmp[0]]))
            tmp = tmp + ('', ', '.join(lista),)
        verb_sheet.append(tmp)

    # usunięcie pustego pierwszego arkusza
    sheet1 = wb['Sheet']
    wb.remove(sheet1)

    wb.save(filename=output_path)


# total: frekwencja rzeczowników i czasowników razem we wszystkich
total_noun_counter.update(Counter(total_nouns))
total_verb_counter.update(Counter(total_verbs))

tmp_total_nouns = list(total_noun_counter.most_common())
tmp_total_verbs = list(total_verb_counter.most_common())

total_path = Path('.').parent / 'output/wyniki/slowa_RAZEM.xlsx'
wb = openpyxl.Workbook()
ws = wb.active

noun_sheet = create_sheet(wb=wb, title="Rzeczowniki", columns=['Rzeczownik', 'Frekwencja', '', 'Słowa'])
verb_sheet = create_sheet(wb=wb, title="Czasowniki", columns=['Czasowniki', 'Frekwencja', '', 'Słowa'])

for tmp in tmp_total_nouns:
    if tmp[0] in total_nouns_text:
        lista = list(set(total_nouns_text[tmp[0]]))
        tmp = tmp + ('', ', '.join(lista),)
    noun_sheet.append(tmp)

for tmp in tmp_total_verbs:
    if tmp[0] in total_verbs_text:
        lista = list(set(total_verbs_text[tmp[0]]))
        tmp = tmp + ('', ', '.join(lista),)
    verb_sheet.append(tmp)

# usunięcie pustego pierwszego arkusza
sheet1 = wb['Sheet']
wb.remove(sheet1)

wb.save(filename=total_path)
