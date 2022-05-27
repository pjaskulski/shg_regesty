""" tworzenie list regestów ze Słownika Historyczno-Geograficznego """

import re
import csv
import sys
from bs4 import BeautifulSoup, NavigableString
from pathlib import Path
import openpyxl
from tools import create_sheet, get_regesty, get_year_from_regest, semisplit
from tools import get_source_from_regest, get_contents_from_regest


# mamy bardzo duże pola i bez zwiększenia limitu skrypt się wykłada
csv.field_size_limit(sys.maxsize)

slowniki = {1: 'Benedyktyni',
            2: 'Chełmno',
            3: 'Kraków',
            4: 'Lublin',
            5: 'Lublin_zaginione',
            6: 'Płock',
            7: 'Poznań',
            8: 'Sanok',
            9: 'Wieluń',
            10: 'Wyszogród',
            11: 'Warszawa',
            12: 'Liw'
}

lista = {
            r'włas[n]?\.\s+szlach\.':'szlachecka',
            r'włas[n]?\.\s+kl\.':'klasztorna',
            r'włas[n]?\.\s+opata':'opata',
            r'włas[n]?\.\s+konw.':'konwentu',
            r'włas[n]?\.\s+ryc\.':'rycerska',
            r'należy\s+do\s+stołu\s+konw\.':'konwentu',
            r'należy\s+do\s+bpa\s+lubus\.':'biskupia',
            r'własn\.\s+podzielona\.':'podzielona',
            r'wieś\s+szlach\.':'szlachecka',
            r'należy\s+do\s+kl\.':'klasztorna',
            r'włas[n]?\.\s+wwdy':'wojewody',
            r'należy\s+do\s+[a-zA-ząśężźćńłó]+\s+stołu\s+opata':'opata',
            r'należy\s+do\s+stołu\s+opata':'opata',
            r'włas[n]?\.\s+rządu':'rządowa',
            r'włas[n]?\.\s+król\.':'królewska',
            r'należy\s+do\s+króla':'królewska',
            r'należy\s+do\s+opactwa':'opata',
            r'włas[n]?\.\s+szlach\.\s+i\s+prep\.':'szlachecka, prepozyta',
            r'włas[n]?\.\s+bpa':'biskupia',
            r'włas[n]?\.\s+szlach\.,\s+następnie\s+kl\.':'szlachecka, klasztorna',
            r'włas[n]?\.\s+bpa[a-zA-Z\.\s]+,\s+następnie\s+rządowa':'biskupia, rządowa',
            r'włas[n]?. kl.[a-zA-Ząśężźźćńłó\.,\s]+pleb\.,\s+[a-zA-Zśężźćńłó]\s+szlach\.':'klasztorna, plebana, szlachecka',
            r'włas[n]?\.\s+kl\.[a-zA-Ząśężźźćńłó\.,\s]+prep\.':'klasztorna, prepozyta',
            r'włas[n]?\.\s+bp':'biskupia',
            r'włas[n]?\.\s+arcbpstwa':'arcybiskupia',
            r'włas[n]?\.\s+książęca':'książęca',
            r'włas[n]?\.\s+książęca[a-zA-Z1-9ąśężźźćńłó\.,\s]+\s+król\.':'książęca, królewska',
            r'włas[n]?\.\s+książęca[a-zA-Ząśężźćńłó,\s]+szlach\./gm':'książęca, szlachecka',
            r'włas[n]?\.\s+książąt':'książęca',
            r'włas[n]?\.\s+książęca[a-zA-Ząśężźćńłó,\s]+kl\.':'książęca, klasztorna'
        }


def all_caps(value: str) -> True:
    """ czy nazwa pisana głównie dużymi literami? """
    result = True
    off = False
    lower_counter = 0
    for ch in value:
        if ch == '[':
            off = True
        elif ch == ']':
            off = False
        elif ch == '–':
            break
        elif not off and ch not in [' ', ',', ';', '.'] and ch.islower():
            lower_counter += 1

    if lower_counter > len(nazwa) - lower_counter:
        result = False

    return result


def remove_html(value: str) -> str:
    """ czyszczenie z pozostałości HTML-a """
    pattern = r'<.*?>'
    value = re.sub(pattern, '', value)
    value = value.replace(r'\n', ' ')
    return value


def process_biblio(sheet, m_id, m_nazwa, point_num, point_text):
    """ przetwarzanie punktu 7 (literatura) """
    reg_list = semisplit(point_text)
    reg_list = [remove_html(reg) for reg in reg_list]
    for item in reg_list:
        biblio = tuple([m_id, m_nazwa, point_num, item])
        sheet.append(biblio)


def process_archeo(sheet, m_id, m_nazwa, point_num, point_text):
    """ przetwarzanie punktu 8 (archeologia i zabytki) """
    reg_list = semisplit(point_text)
    reg_list = [remove_html(reg) for reg in reg_list]
    for item in reg_list:
        archeo = tuple([m_id, m_nazwa, point_num, item.strip()])
        sheet.append(archeo)


def process_header(sheet, m_id, m_nazwa, point_num, point_text) -> int:
    """ przetwarzanie punktu 0 (nagłówek, obecnie tylko podział na akapity) """
    global regesty_count
    regest_num = 0
    reg_list = re.split('<p>', point_text)
    reg_list = [remove_html(reg) for reg in reg_list]
    for item in reg_list:
        # jeżeli możliwy podział na regesty
        if ');' in item:
            reg_list, ownership = get_regesty(item)
            reg_list = [remove_html(reg) for reg in reg_list]
            if reg_list:
                for reg_item in reg_list:
                    regest = tuple([m_id, m_nazwa, point_num, reg_item.strip()])
                    sheet.append(regest)
                    # wyłączenie sumowania dla p. 0 regesty_count[point_num] += 1
                    # regest_num += 1
        else:
            akapit = tuple([m_id, m_nazwa, point_num, item.strip()])
            sheet.append(akapit)

    return regest_num


def process_point(sheet, m_id, m_nazwa, point_num, point_text) ->int:
    """przetwarzanie treści punktu z zapisem do wskazanego arkusza
       sheet - wskazanie na arkusz
       m_id - identyfikator miejscowości
       nazwa - nazwa miejscowości
       point_num - numer punktu
       point_text - treść punktu
    """
    global regesty_count
    global regesty_slownik_count
    global place_owner

    reg_list = []
    point = point_num

    # if 'Własn. szlach. w kluczu melsztyńskim' in point_text:
    #     print()

    reg_list, ownership = get_regesty(point_text)

    reg_list = [remove_html(reg) for reg in reg_list]

    if reg_list:
        # pierwszy regest może wymagać podzielenia
        pattern = r'\.\s{1}\d{4}'
        pattern2 = r'\s+r\.\s{1}\d{4}'
        match = re.search(pattern, reg_list[0])
        match2 = re.search(pattern2, reg_list[0])
        if match and not match2:
            pos = reg_list[0].find('. 1')
            if pos != -1:
                part1 = reg_list[0][:pos + 1]
                part2 = reg_list[0][pos + 2:]
                reg_list[0] = part2
                reg_list.insert(0, part1)

    # własność
    if point_num == 3:
        for reg in reg_list:
            year = get_year_from_regest(reg)
            for pattern, owner_type in lista.items():
                match = re.search(pattern, reg.lower())
                if match:
                    if (m_id, year) in place_owner:
                        _m_id, _year, owner_types = place_owner[(m_id, year)]
                        if owner_type not in owner_types:
                            if not owner_types.endswith(','):
                                owner_types += ', '
                            owner_types += owner_type
                            place_owner[(m_id, year)] = (m_nazwa, year, owner_types)
                    else:
                        place_owner[(m_id, year)] = (m_nazwa, year, owner_type)

    # specjalna obsługa dla p. 1
    if point_num == 1:
        new_reg_list = []
        for item in reg_list:
            if ")," in item:
                tmp = item.split("),")
                tmp = [r+')' if not r.endswith(")") else r for r in tmp]
                new_reg_list = new_reg_list + tmp
            elif ")." in item:
                tmp = item.split(").")
                tmp = [r+')' if not r.endswith(")") else r for r in tmp]
                new_reg_list = new_reg_list + tmp
            else:
                new_reg_list.append(item)
        reg_list = new_reg_list

    prev_year = ''
    regest_num = 0
    for item in reg_list:
        # if 'n. par. Nowa Słupia (DLb. II 490) [' in item:
        #     print()
        item = item.strip()
        year = get_year_from_regest(item)
        # jeżeli regest nie ma roku to przyjmujemy poprzedni
        if year.strip() == '' and regest_num > 0 and prev_year != '':
            year = prev_year + '!'
        else:
            prev_year = year
        source = get_source_from_regest(item)
        contents = get_contents_from_regest(item, year, source, m_nazwa)
        contents = contents.replace('[@', '[').replace('@]',']')
        item = item.replace('[@', '[').replace('@]',']')
        tmp_regest = [m_id, m_nazwa, point, year, contents, source, item]
        regest = tuple(tmp_regest)
        sheet.append(regest)
        regesty_count[point_num] += 1
        regest_num += 1

    return regest_num


# -------------- Główny program ------------------------------------------------
if __name__ == '__main__':
    regesty_count = {0:0, 1:0, 2:0, 3:0, 4:0, 5:0, 6:0}

    regesty_slownik_count = {0:0, 1:0, 2:0, 3:0, 4:0, 5:0, 6:0, 7:0, 8:0, 9:0, 10:0, 11:0, 12:0}

    for nr in range(1, 13):
        print(f"Słownik: {slowniki[nr]} ({nr})")

        place_owner = {}

        output_path = Path('.').parent / f'output/regesty_{slowniki[nr]}.xlsx'
        input_path = Path('.').parent / f'output/hasla_{nr}_pkt.csv'

        wb = openpyxl.Workbook()
        ws = wb.active

        # nagłówek, tylko podział wg akapitów
        point_sheet_columns = ['Miejscowosc_Id', 'Nazwa', 'Punkt', 'Akapit']
        regest_sheet_0 = create_sheet(wb=wb, title="Nagłówek, p. 0", columns=point_sheet_columns)

        point_sheet_columns = ['Miejscowość_Id', 'Nazwa', 'Punkt', 'Rok',
                               'Treść', 'Źródła', 'Cały Regest']

        regest_sheet_1 = create_sheet(wb=wb, title="Regesty p. 1", columns=point_sheet_columns)
        regest_sheet_2 = create_sheet(wb=wb, title="Regesty p. 2", columns=point_sheet_columns)
        regest_sheet_2a = create_sheet(wb=wb, title="Regesty p. 2a", columns=point_sheet_columns)
        regest_sheet_2b = create_sheet(wb=wb, title="Regesty p. 2b", columns=point_sheet_columns)
        regest_sheet_2c = create_sheet(wb=wb, title="Regesty p. 2c", columns=point_sheet_columns)
        regest_sheet_2d = create_sheet(wb=wb, title="Regesty p. 2d", columns=point_sheet_columns)
        regest_sheet_2e = create_sheet(wb=wb, title="Regesty p. 2e", columns=point_sheet_columns)
        regest_sheet_2f = create_sheet(wb=wb, title="Regesty p. 2f", columns=point_sheet_columns)

        regest_sheet_3 = create_sheet(wb=wb, title="Regesty p. 3", columns=point_sheet_columns)
        regest_sheet_3a = create_sheet(wb=wb, title="Regesty p. 3a", columns=point_sheet_columns)
        regest_sheet_3b = create_sheet(wb=wb, title="Regesty p. 3b", columns=point_sheet_columns)
        regest_sheet_3c = create_sheet(wb=wb, title="Regesty p. 3c", columns=point_sheet_columns)
        regest_sheet_3d = create_sheet(wb=wb, title="Regesty p. 3d", columns=point_sheet_columns)
        regest_sheet_3e = create_sheet(wb=wb, title="Regesty p. 3e", columns=point_sheet_columns)
        regest_sheet_3f = create_sheet(wb=wb, title="Regesty p. 3f", columns=point_sheet_columns)
        
        regest_sheet_4 = create_sheet(wb=wb, title="Regesty p. 4", columns=point_sheet_columns)
        regest_sheet_4a = create_sheet(wb=wb, title="Regesty p. 4a", columns=point_sheet_columns)
        regest_sheet_4b = create_sheet(wb=wb, title="Regesty p. 4b", columns=point_sheet_columns)
        regest_sheet_4c = create_sheet(wb=wb, title="Regesty p. 4c", columns=point_sheet_columns)
        regest_sheet_4d = create_sheet(wb=wb, title="Regesty p. 4d", columns=point_sheet_columns)
        regest_sheet_4e = create_sheet(wb=wb, title="Regesty p. 4e", columns=point_sheet_columns)
        regest_sheet_4f = create_sheet(wb=wb, title="Regesty p. 4f", columns=point_sheet_columns)

        regest_sheet_5 = create_sheet(wb=wb, title="Regesty p. 5", columns=point_sheet_columns)
        regest_sheet_5a = create_sheet(wb=wb, title="Regesty p. 5a", columns=point_sheet_columns)
        regest_sheet_5b = create_sheet(wb=wb, title="Regesty p. 5b", columns=point_sheet_columns)
        regest_sheet_5c = create_sheet(wb=wb, title="Regesty p. 5c", columns=point_sheet_columns)
        regest_sheet_5d = create_sheet(wb=wb, title="Regesty p. 5d", columns=point_sheet_columns)
        regest_sheet_5e = create_sheet(wb=wb, title="Regesty p. 5e", columns=point_sheet_columns)
        regest_sheet_5f = create_sheet(wb=wb, title="Regesty p. 5f", columns=point_sheet_columns)

        regest_sheet_6 = create_sheet(wb=wb, title="Regesty p. 6", columns=point_sheet_columns)
        regest_sheet_6a = create_sheet(wb=wb, title="Regesty p. 6a", columns=point_sheet_columns)
        regest_sheet_6b = create_sheet(wb=wb, title="Regesty p. 6b", columns=point_sheet_columns)
        regest_sheet_6c = create_sheet(wb=wb, title="Regesty p. 6c", columns=point_sheet_columns)
        regest_sheet_6d = create_sheet(wb=wb, title="Regesty p. 6d", columns=point_sheet_columns)
        regest_sheet_6e = create_sheet(wb=wb, title="Regesty p. 6e", columns=point_sheet_columns)
        regest_sheet_6f = create_sheet(wb=wb, title="Regesty p. 6f", columns=point_sheet_columns)

        # bibliografia, tylko podział po średnikach?
        point_sheet_columns = ['Miejscowosc_Id', 'Nazwa', 'Punkt', 'Publikacja']
        regest_sheet_7 = create_sheet(wb=wb, title="Publikacje p. 7", columns=point_sheet_columns)

        # archeologia i zabytki, tylko podział po średnikach?
        point_sheet_columns = ['Miejscowosc_Id', 'Nazwa', 'Punkt', 'Obiekt']
        regest_sheet_8 = create_sheet(wb=wb, title="Archeologia, zabytki p. 8", columns=point_sheet_columns)

        # przetwarzanie wierszy
        with open(input_path, mode='r', encoding='utf-8') as csv_file:
            csv_reader = csv.DictReader(csv_file, delimiter='#')
            #lines = len(list(csv_reader))
            licznik = 0
            for row in csv_reader:
                p0 = p1= p2 = p3 = p4 = p5 = p6 = p7 = p8 = p9 = ''
                licznik += 1
                print(f'Przetwarzam wiersz {licznik}.')
                miejsce_id = row['id'].strip()
                nazwa = row["header"]
                nazwa = re.sub(r'<.*?>', '', nazwa)

                # odnośniki pomijamy
                if not all_caps(nazwa):
                    continue

                p0 = row['0'].strip()
                p1 = row['1'].strip()
                p2 = row['2'].strip()
                p3 = row['3'].strip()
                p4 = row['4'].strip()
                p5 = row['5'].strip()
                p6 = row['6'].strip()
                p7 = row['7'].strip()
                p8 = row['8'].strip()
                p9 = row['9'].strip()

                # Nagłówek z podziałem na paragrafy (p. 0)
                #if p0:
                #    process_header(regest_sheet_0, miejsce_id, nazwa, 0, p0)

                if p0:
                    liczba = process_header(regest_sheet_0, miejsce_id, nazwa, 0, p0)
                    #regesty_slownik_count[nr] += liczba
                # Punkt 1 bez podpunktów
                if p1:
                    liczba = process_point(regest_sheet_1, miejsce_id, nazwa, 1, p1)
                    regesty_slownik_count[nr] += liczba
                # Punkt 2 i dalsze z możliwymi podpunktami
                if p2:
                    pattern = r'{.*?}'
                    if re.search(pattern, p2):
                        p_part = {}
                        parts = re.split(pattern, p2)
                        symbols = [remove_html(sym.group()) for sym in re.finditer('(' + pattern + ')', p2)]
                        if parts[0].strip() != '':
                            p_part['2'] = parts[0]
                        for i in range(1, len(parts)):
                            p_part[symbols[i-1]] = parts[i]
                        
                        for key, value in p_part.items():
                            key = key.replace('.','').replace('-','').lower()
                            if key == '2':
                                liczba = process_point(regest_sheet_2, miejsce_id, nazwa, 2, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'a' in key:
                                liczba = process_point(regest_sheet_2a, miejsce_id, nazwa, 2, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'b' in key:
                                liczba = process_point(regest_sheet_2b, miejsce_id, nazwa, 2, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'c' in key:
                                liczba = process_point(regest_sheet_2c, miejsce_id, nazwa, 2, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'd' in key:
                                liczba = process_point(regest_sheet_2d, miejsce_id, nazwa, 2, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'e' in key:
                                liczba = process_point(regest_sheet_2e, miejsce_id, nazwa, 2, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'f' in key:
                                liczba = process_point(regest_sheet_2f, miejsce_id, nazwa, 2, value)
                                regesty_slownik_count[nr] += liczba
                    else:
                        liczba = process_point(regest_sheet_2, miejsce_id, nazwa, 2, p2)
                        regesty_slownik_count[nr] += liczba
                if p3:
                    pattern = r'{.*?}'
                    if re.search(pattern, p3):
                        p_part = {}
                        parts = re.split(pattern, p3)
                        symbols = [remove_html(sym.group()) for sym in re.finditer('(' + pattern + ')', p3)]
                        if parts[0].strip() != '':
                            p_part['3'] = parts[0]
                        for i in range(1, len(parts)):
                            p_part[symbols[i-1]] = parts[i]
                        
                        for key, value in p_part.items():
                            key = key.replace('.','').replace('-','').lower()
                            if key == '3':
                                liczba = process_point(regest_sheet_3, miejsce_id, nazwa, 3, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'a' in key:
                                liczba = process_point(regest_sheet_3a, miejsce_id, nazwa, 3, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'b' in key:
                                liczba = process_point(regest_sheet_3b, miejsce_id, nazwa, 3, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'c' in key:
                                liczba = process_point(regest_sheet_3c, miejsce_id, nazwa, 3, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'd' in key:
                                liczba = process_point(regest_sheet_3d, miejsce_id, nazwa, 3, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'e' in key:
                                liczba = process_point(regest_sheet_3e, miejsce_id, nazwa, 3, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'f' in key:
                                liczba = process_point(regest_sheet_3f, miejsce_id, nazwa, 3, value)
                                regesty_slownik_count[nr] += liczba
                    else:
                        liczba = process_point(regest_sheet_3, miejsce_id, nazwa, 3, p3)
                        regesty_slownik_count[nr] += liczba
                if p4:
                    pattern = r'{.*?}'
                    if re.search(pattern, p4):
                        p_part = {}
                        parts = re.split(pattern, p4)
                        symbols = [remove_html(sym.group()) for sym in re.finditer('(' + pattern + ')', p4)]
                        if parts[0].strip() != '':
                            p_part['4'] = parts[0]
                        for i in range(1, len(parts)):
                            p_part[symbols[i-1]] = parts[i]
                        
                        for key, value in p_part.items():
                            key = key.replace('.','').replace('-','').lower()
                            if key == '4':
                                liczba = process_point(regest_sheet_4, miejsce_id, nazwa, 4, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'a' in key:
                                liczba = process_point(regest_sheet_4a, miejsce_id, nazwa, 4, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'b' in key:
                                liczba = process_point(regest_sheet_4b, miejsce_id, nazwa, 4, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'c' in key:
                                liczba = process_point(regest_sheet_4c, miejsce_id, nazwa, 4, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'd' in key:
                                liczba = process_point(regest_sheet_4d, miejsce_id, nazwa, 4, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'e' in key:
                                liczba = process_point(regest_sheet_4e, miejsce_id, nazwa, 4, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'f' in key:
                                liczba = process_point(regest_sheet_4f, miejsce_id, nazwa, 4, value)
                                regesty_slownik_count[nr] += liczba
                    else:
                        liczba = process_point(regest_sheet_4, miejsce_id, nazwa, 4, p4)
                        regesty_slownik_count[nr] += liczba
                if p5:
                    pattern = r'{.*?}'
                    if re.search(pattern, p5):
                        p_part = {}
                        parts = re.split(pattern, p5)
                        symbols = [remove_html(sym.group()) for sym in re.finditer('(' + pattern + ')', p5)]
                        if parts[0].strip() != '':
                            p_part['5'] = parts[0]
                        for i in range(1, len(parts)):
                            p_part[symbols[i-1]] = parts[i]

                        for key, value in p_part.items():
                            key = key.replace('.','').replace('-','').lower()
                            if key == '5':
                                liczba = process_point(regest_sheet_5, miejsce_id, nazwa, 5, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'a' in key:
                                liczba = process_point(regest_sheet_5a, miejsce_id, nazwa, 5, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'b' in key:
                                liczba = process_point(regest_sheet_5b, miejsce_id, nazwa, 5, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'c' in key:
                                liczba = process_point(regest_sheet_5c, miejsce_id, nazwa, 5, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'd' in key:
                                liczba = process_point(regest_sheet_5d, miejsce_id, nazwa, 5, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'e' in key:
                                liczba = process_point(regest_sheet_5e, miejsce_id, nazwa, 5, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'f' in key:
                                liczba = process_point(regest_sheet_5f, miejsce_id, nazwa, 5, value)
                                regesty_slownik_count[nr] += liczba
                    else:
                        liczba = process_point(regest_sheet_5, miejsce_id, nazwa, 5, p5)
                        regesty_slownik_count[nr] += liczba
                if p6:
                    pattern = r'{.*?}'
                    if re.search(pattern, p6):
                        p_part = {}
                        parts = re.split(pattern, p6)
                        symbols = [remove_html(sym.group()) for sym in re.finditer('(' + pattern + ')', p6)]
                        if parts[0].strip() != '':
                            p_part['6'] = parts[0]
                        for i in range(1, len(parts)):
                            p_part[symbols[i-1]] = parts[i]
                        
                        for key, value in p_part.items():
                            key = key.replace('.','').replace('-','').lower()
                            if key == '6':
                                liczba = process_point(regest_sheet_6, miejsce_id, nazwa, 6, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'a' in key:
                                liczba = process_point(regest_sheet_6a, miejsce_id, nazwa, 6, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'b' in key:
                                liczba = process_point(regest_sheet_6b, miejsce_id, nazwa, 6, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'c' in key:
                                liczba = process_point(regest_sheet_6c, miejsce_id, nazwa, 6, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'd' in key:
                                liczba = process_point(regest_sheet_6d, miejsce_id, nazwa, 6, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'e' in key:
                                liczba = process_point(regest_sheet_6e, miejsce_id, nazwa, 6, value)
                                regesty_slownik_count[nr] += liczba
                            elif 'f' in key:
                                liczba = process_point(regest_sheet_6f, miejsce_id, nazwa, 6, value)
                                regesty_slownik_count[nr] += liczba
                    else:
                        liczba = process_point(regest_sheet_6, miejsce_id, nazwa, 6, p6)
                        regesty_slownik_count[nr] += liczba
                # literatura
                if p7:
                    process_biblio(regest_sheet_7, miejsce_id, nazwa, 7, p7)
                # literatura
                if p8:
                    process_archeo(regest_sheet_8, miejsce_id, nazwa, 8, p8)

        # usunięcie pustego pierwszego arkusza
        sheet1 = wb['Sheet']
        wb.remove(sheet1)
        # jeżeli są puste arkusze to należy je usunąć
        if regest_sheet_0.max_row == 1:
            wb.remove(regest_sheet_0)

        if regest_sheet_1.max_row == 1:
            wb.remove(regest_sheet_1)

        if regest_sheet_2.max_row == 1:
            wb.remove(regest_sheet_2)
        if regest_sheet_2a.max_row == 1:
            wb.remove(regest_sheet_2a)
        if regest_sheet_2b.max_row == 1:
            wb.remove(regest_sheet_2b)
        if regest_sheet_2c.max_row == 1:
            wb.remove(regest_sheet_2c)
        if regest_sheet_2d.max_row == 1:
            wb.remove(regest_sheet_2d)
        if regest_sheet_2e.max_row == 1:
            wb.remove(regest_sheet_2e)
        if regest_sheet_2f.max_row == 1:
            wb.remove(regest_sheet_2f)

        if regest_sheet_3.max_row == 1:
            wb.remove(regest_sheet_3)
        if regest_sheet_3a.max_row == 1:
            wb.remove(regest_sheet_3a)
        if regest_sheet_3b.max_row == 1:
            wb.remove(regest_sheet_3b)
        if regest_sheet_3c.max_row == 1:
            wb.remove(regest_sheet_3c)
        if regest_sheet_3d.max_row == 1:
            wb.remove(regest_sheet_3d)
        if regest_sheet_3e.max_row == 1:
            wb.remove(regest_sheet_3e)
        if regest_sheet_3f.max_row == 1:
            wb.remove(regest_sheet_3f)

        if regest_sheet_4.max_row == 1:
            wb.remove(regest_sheet_4)
        if regest_sheet_4a.max_row == 1:
            wb.remove(regest_sheet_4a)
        if regest_sheet_4b.max_row == 1:
            wb.remove(regest_sheet_4b)
        if regest_sheet_4c.max_row == 1:
            wb.remove(regest_sheet_4c)
        if regest_sheet_4d.max_row == 1:
            wb.remove(regest_sheet_4d)
        if regest_sheet_4e.max_row == 1:
            wb.remove(regest_sheet_4e)
        if regest_sheet_4f.max_row == 1:
            wb.remove(regest_sheet_4f)

        if regest_sheet_5.max_row == 1:
            wb.remove(regest_sheet_5)
        if regest_sheet_5a.max_row == 1:
            wb.remove(regest_sheet_5a)
        if regest_sheet_5b.max_row == 1:
            wb.remove(regest_sheet_5b)
        if regest_sheet_5c.max_row == 1:
            wb.remove(regest_sheet_5c)
        if regest_sheet_5d.max_row == 1:
            wb.remove(regest_sheet_5d)
        if regest_sheet_5e.max_row == 1:
            wb.remove(regest_sheet_5e)
        if regest_sheet_5f.max_row == 1:
            wb.remove(regest_sheet_5f)

        if regest_sheet_6.max_row == 1:
            wb.remove(regest_sheet_6)
        if regest_sheet_6a.max_row == 1:
            wb.remove(regest_sheet_6a)
        if regest_sheet_6b.max_row == 1:
            wb.remove(regest_sheet_6b)
        if regest_sheet_6c.max_row == 1:
            wb.remove(regest_sheet_6c)
        if regest_sheet_6d.max_row == 1:
            wb.remove(regest_sheet_6d)
        if regest_sheet_6e.max_row == 1:
            wb.remove(regest_sheet_6e)
        if regest_sheet_6f.max_row == 1:
            wb.remove(regest_sheet_6f)

        if regest_sheet_7.max_row == 1:
            wb.remove(regest_sheet_7)

        if regest_sheet_8.max_row == 1:
            wb.remove(regest_sheet_8)

        # arkusz z własnością
        wlasnosc_sheet = create_sheet(wb=wb, title="Własność", columns=['Miejsc_ID', 'Miejscowość', 'Rok', 'Własność'])
        for key_tuple, owner_lista in place_owner.items():
            miejscowosc_id, _year = key_tuple
            miejscowosc, rok, owner_types = owner_lista

            # bez powtórzeń typów własności
            tmp = owner_types.split(',')
            if len(tmp) > 1:
                tmp = [t.strip() for t in tmp]
                tmp = list(set(tmp))
                owner_types = ', '.join(tmp)

            wiersz = (miejscowosc_id, miejscowosc, rok, owner_types)
            wlasnosc_sheet.append(wiersz)

        # zapis pliku xlsx
        wb.save(filename=output_path)

    suma = 0
    for i in range(1, 7):
        print(f"p.{i}", regesty_count[i])
        suma += regesty_count[i]
    print(f"razem: {suma}")
    print()
    suma = 0
    for nr in range(1, 13):
        print(f"Słownik {slowniki[nr]}: {regesty_slownik_count[nr]}")
        suma += regesty_slownik_count[nr]
    print(f"razem: {suma}")